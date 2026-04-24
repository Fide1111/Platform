# New Claim Register - Update the columns names, copy and paste from the previous one to the latest one

import shutil
import sqlite3
import pandas as pd 
import os
import numpy as np
from datetime import datetime
from openpyxl.utils.dataframe import dataframe_to_rows # type: ignore
from sqlalchemy import create_engine # type: ignore
from openpyxl import load_workbook  # type: ignore
from IPython.display import display  # type: ignore
pd.set_option('display.max_rows', 500)
timestamp = datetime.now().strftime("%Y-%m-%d")
years = range(2010,2027)

list_of_lob = {
    2111: "EC",
    2121: "Individual_PA",
    2122: "Group_PA_excl_Student_PA",
    2123: "Travel",
    222: "Hospital_Cash_Plan",
    223: "Group_Medical",
    311: "Fire_Insurance",
    323: "Burglary",
    3261: "Mhinery",
    3262: "Electronic_Equipment",
    3263: "Contractor_Plant",
    32711: "Household",
    3273: "PAR",
    3283: "CAR",
    41: "Motor_Damage_to_Vehicle",
    43: "Motor_TPL", 
    44: "Motor_Passenger_Liability",
    62: "Aviation",
    71: "Marine_Exh_All_Risks",
    72: "Inland_Transit_Insurance",
    79: "Marine_Cargo_Exh_All_Risks",
    82: "Prof_Liab_excl_Lawyers_Liab",
    83: "Cyber_Insurance",
    854: "Public_Liability",
    92: "Fidelity_Guarantee",
}

TEMPLATE_COLUMNS_DATA_RECORD = [
    "AG Classification", "Biz Source", "Op Year", "Claim No.", "Type Sap", "Policy No.", "Class", 
    "Date of Occurrence", "Reported Date", "Accident yr", 
    "Claim Reserves \n2010", "Claim Reserves \n2011", "Claim Reserves \n2012", "Claim Reserves \n2013", 
    "Claim Reserves \n2014", "Claim Reserves \n2015", "Claim Reserves \n2016", "Claim Reserves \n2017", 
    "Claim Reserves \n2018", "Claim Reserves \n2019", "Claim Reserves \n2020", "Claim Reserves \n2021", 
    "Claim Reserves \n2022", "Claim Reserves \n2023", "Claim Reserves \n2024", "Claim Reserves \n2025", "Claim Reserves \n2026",
    "Settled Amount \n2010", "Settled Amount \n2011", "Settled Amount \n2012", "Settled Amount \n2013", 
    "Settled Amount \n2014", "Settled Amount \n2015", "Settled Amount \n2016", "Settled Amount \n2017", 
    "Settled Amount \n2018", "Settled Amount \n2019", "Settled Amount \n2020", "Settled Amount \n2021", 
    "Settled Amount \n2022", "Settled Amount \n2023", "Settled Amount \n2024", "Settled Amount \n2025", "Settled Amount \n2026",
    "Excess \n2010", "Excess \n2011", "Excess \n2012", "Excess \n2013", "Excess \n2014", "Excess \n2015", 
    "Excess \n2016", "Excess \n2017", "Excess \n2018", "Excess \n2019", "Excess \n2020", "Excess \n2021", 
    "Excess \n2022", "Excess \n2023", "Excess \n2024", "Excess \n2025", "Excess \n2026",
    "RI Claims Reserves \n2010", "RI Claims Reserves \n2011", "RI Claims Reserves \n2012", "RI Claims Reserves \n2013", 
    "RI Claims Reserves \n2014", "RI Claims Reserves \n2015", "RI Claims Reserves \n2016", "RI Claims Reserves \n2017", 
    "RI Claims Reserves \n2018", "RI Claims Reserves \n2019", "RI Claims Reserves \n2020", "RI Claims Reserves \n2021", 
    "RI Claims Reserves \n2022", "RI Claims Reserves \n2023", "RI Claims Reserves \n2024", "RI Claims Reserves \n2025", "RI Claims Reserves \n2026",
    "RI Settled Amount \n2010", "RI Settled Amount \n2011", "RI Settled Amount \n2012", "RI Settled Amount \n2013", 
    "RI Settled Amount \n2014", "RI Settled Amount \n2015", "RI Settled Amount \n2016", "RI Settled Amount \n2017", 
    "RI Settled Amount \n2018", "RI Settled Amount \n2019", "RI Settled Amount \n2020", "RI Settled Amount \n2021", 
    "RI Settled Amount \n2022", "RI Settled Amount \n2023", "RI Settled Amount \n2024", "RI Settled Amount \n2025", "RI Settled Amount \n2026",
    "Gr Clm Incurred \n2010", "Gr Clm Incurred \n2011", "Gr Clm Incurred \n2012", "Gr Clm Incurred \n2013", 
    "Gr Clm Incurred \n2014", "Gr Clm Incurred \n2015", "Gr Clm Incurred \n2016", "Gr Clm Incurred \n2017", 
    "Gr Clm Incurred \n2018", "Gr Clm Incurred \n2019", "Gr Clm Incurred \n2020", "Gr Clm Incurred \n2021", 
    "Gr Clm Incurred \n2022", "Gr Clm Incurred \n2023", "Gr Clm Incurred \n2024", "Gr Clm Incurred \n2025", "Gr Clm Incurred \n2026",
    "Gr Clm Incurred Accumulated",
    "RI Clm Incurred \n2010", "RI Clm Incurred \n2011", "RI Clm Incurred \n2012", "RI Clm Incurred \n2013", 
    "RI Clm Incurred \n2014", "RI Clm Incurred \n2015", "RI Clm Incurred \n2016", "RI Clm Incurred \n2017", 
    "RI Clm Incurred \n2018", "RI Clm Incurred \n2019", "RI Clm Incurred \n2020", "RI Clm Incurred \n2021", 
    "RI Clm Incurred \n2022", "RI Clm Incurred \n2023", "RI Clm Incurred \n2024", "RI Clm Incurred \n2025", "RI Clm Incurred \n2026",
    "RI Clm Incurred Accumulated",
    "Net Clm Incurred 2020", "Net Clm Incurred 2021", "Net Clm Incurred 2022", "Net Clm Incurred 2023", 
    "Net Clm Incurred 2024", "Net Clm Incurred 2025", "Net Clm Incurred 2026", "Net Clm Incurred Accumulated",
    "Valid case #", "Product", "Status", "Gr clm > 500k?", "Gr clm > 1m?", "Event", "Check seq with Register"
]

# TEMPLATE_COLUMNS_CLAIMS_REGISTER_DIRECT = [
#     "Year",
#     "Claims NO.",
#     "Type Sap", 
#     "Policy NO.",
#     "Class",
#     "Date of Occurrence",
#     "Reported Date",
#     "Occurred and reported in same year",
#     "Claim Reserves 2026",
#     "Voucher NO. 2026",
#     "Settled Amount 2026",
#     "Ref. 2026",
#     "Payment Method 2026",
#     "Excess 2026",
#     "Ref. NO. 2026",
#     "Payment Method",  # Note: this seems duplicated/misaligned in original
#     "Claim Reserves 2025",
#     "Voucher NO. 2025",
#     "Settled Amount 2025",
#     "Ref. 2025",
#     "Payment Method 2025",
#     "Excess 2025",
#     "Ref. NO. 2025",
#     "Payment Method",  # duplicated
#     "Claim Reserves 2024",
#     "Voucher NO. 2024",
#     "Settled Amount 2024",
#     "Ref. 2024",
#     "Payment Method 2024",
#     "Excess 2024",
#     "Ref. NO. 2024",
#     "Payment Method",  # duplicated
#     "Claim Reserves 2023",
#     "Voucher NO. 2023",
#     "Settled Amount 2023",
#     "Ref. 2023",
#     "Payment Method 2023",
#     "Excess 2023",
#     "Ref. NO. 2023",
#     "Payment Method",  # duplicated
#     "Claim Reserves 2022",
#     "Voucher NO. 2022",
#     "Settled Amount 2022",
#     "Ref. 2022",
#     "Excess 2022",
#     "Ref. NO. 2022",
#     "Claim Reserves 2021",
#     "Voucher NO. 2021",
#     "Settled Amount 2021",
#     "Ref. 2021",
#     "Excess 2021",
#     "Ref. NO. 2021",
#     "Claim Reserves 2020",
#     "Voucher NO. 2020",
#     "Settled Amount 2020",
#     "Ref. 2020",
#     "Excess 2020",
#     "Ref. NO. 2020",
#     "Claim Reserves 2019",
#     "Voucher NO. 2019",
#     "Settled Amount 2019",
#     "Ref. 2019",
#     "Excess 2019",
#     "Ref. NO. 2019",
#     "Claim Reserves 2018",
#     "Voucher NO. 2018",
#     "Settled Amount 2018",
#     "Ref. 2018",
#     "Ref. NO. 2018",
#     "Claim Reserves 2017",
#     "Voucher NO. 2017",
#     "Settled Amount 2017",
#     "Ref. 2017",
#     "Excess 2017",
#     "Ref. NO. 2017",
#     "Voucher NO. 2016",
#     "Claim Reserves 2016",
#     "Settled Amount 2016",
#     "Ref. 2016",
#     "Excess 2016",
#     "Ref. NO. 2016",
#     "Voucher NO. 2015",
#     "Claim Reserves Dec 2015",
#     "Settled Amount Oct - Dec 2015",
#     "Ref. Oct - Dec 2015",
#     "Excess Oct - Dec 2015",
#     "Ref. NO. Oct - Dec 2015",
#     "Claim Reserves 2015 Sept",
#     "Voucher NO. 2015",
#     "Settled Amount 2015 (Jan-Sept)",
#     "Ref. NO. 2015",
#     "Excess 2015 (Jan-Sept)",
#     "Claim Reserves 2014",
#     "Voucher NO. 2014 (Reserve)",
#     "Payment 2014",
#     "Voucher NO. 2014 (Payment)",
#     "Excess 2014",
#     "Claim Reserves 2013",
#     "Voucher NO. 2013 (Reserve)",
#     "Settled Amount 2013",
#     "Voucher NO. 2013 (Payment)",
#     "Excess 2013",
#     "Claim Reserves 2012",
#     "Voucher NO. 2012 (Reserve)",
#     "Payment 2012",
#     "Voucher NO. 2012 (Payment)",
#     "Excess 2012",
#     "Claim Reserves 2011",
#     "Voucher NO. 2011 (Reserve)",
#     "Payment 2011",
#     "Voucher NO. 2011 (Payment)",
#     "Excess 2011",
#     "Claim Reserves 2010",
#     "Voucher NO. 2010 (Reserve)",
#     "Settled Amount 2010",
#     "Voucher NO. 2010 (Payment)",
#     "Excess 2010",
#     # Reinsurer columns
#     "",
#     "Reinsurer",
#     "Rein. Part. In Claims Reserves 2026",
#     "Voucher NO. 2026",
#     "RI Settled Amount 2026",
#     "Ref. 2026",
#     "Rein. Part. In Claims Reserves 2025",
#     "Voucher NO. 2025",
#     "RI Settled Amount 2025",
#     "Ref. 2025",
#     "Rein. Part. In Claims Reserves 2024",
#     "Voucher NO. 2024",
#     "RI Settled Amount 2024",
#     "Ref. 2024",
#     "Rein. Part. In Claims Reserves 2023",
#     "Voucher NO. 2023",
#     "RI Settled Amount 2023",
#     "Ref. 2023",
#     "Rein. Part. In Claims Reserves 2022",
#     "Voucher NO. 2022",
#     "RI Settled Amount 2022",
#     "Ref. 2022",
#     "Rein. Part. In Claims Reserves 2021",
#     "Voucher NO. 2021",
#     "RI Settled Amount 2021",
#     "Ref. 2021",
#     "Rein. Part. In Claims Reserves 2020",
#     "Voucher NO. 2020",
#     "RI Settled Amount 2020",
#     "Ref. 2020",
#     "Rein. Part. In Claims Reserves 2019",
#     "Voucher NO. 2019",
#     "RI Settled Amount 2019",
#     "Ref. 2019",
#     "Rein. Part. In Claims Reserves 2018",
#     "Voucher NO. 2018",
#     "RI Settled Amount 2018",
#     "Ref. 2018",
#     "Rein. Part. In Claims Reserves 2017",
#     "Voucher NO. 2017",
#     "RI Settled Amount 2017",
#     "Ref. 2017",
#     "Voucher NO. 2016",
#     "RI Settled Amount 2016",
#     "Ref. 2016",
#     "Voucher NO. 2015",
#     "RI Settled Amount Oct - Dec 2015",
#     "Ref. Oct - Dec 2015",
#     "RI Settled Amount 2015",
#     "Ref. 2015",
#     "Voucher NO. 2014",
#     "RI Settled Amount 2014",
#     "Ref. 2014",
#     "Voucher NO. 2013",
#     "RI Settled Amount 2013",
#     "Ref. 2013",
#     "Voucher NO. 2012",
#     "Voucher NO. 2011",
#     "Voucher NO. (before 2011)",
#     "RI Settled Amount 2011",
#     "Ref. 2011",
#     "RI Settled Amount 2012",
#     "Ref. 2012",
#     "Settled Amount (before 2011)",
#     "Ref. (before 2011)",
#     # Final summary columns
#     "Salvage", 
#     "Type",
#     "YearOc",
#     "YearRp",
#     "END_AC",
#     "Reserves",
#     "Payments",
#     "Excess",
#     "Salvage",
#     "Ceded Reserves",
#     "Ceded Payments",
#     "",
#     "Acc Res",
#     "Total paid",
#     "Acc excess",
#     "Ceded Res",
#     "Ceded pay-Excess",
#     "",
#     "retention pay",
#     "",
#     "DIFF",
#     "", "", 
#     "JV03-091",
#     "#N/A",
#     "", 
#     "/JV03-091",
#     "", "", 
#     "xol"
# ]
# TEMPLATE_COLUMNS_CLAIMS_REGISTER_FAC_IN = [
#     "Year",
#     "Claims No.",
#     "Type Sap", 
#     "Policy No.",
#     "Class",
#     "Date of Occurrence",
#     "Reported Date",
#     "Occurred and reported in same year",
#     "Claim Reserves 2026",
#     "Voucher NO. 2026",
#     "Settled Amount 2026",
#     "Ref. 2026",
#     "Payment Method 2026",
#     "Excess 2026",
#     "Ref. NO. 2026",
#     "Payment Method",  # Note: this seems duplicated/misaligned in original
#     "Claim Reserves 2025",
#     "Voucher NO. 2025",
#     "Settled Amount 2025",
#     "Ref. 2025",
#     "Payment Method 2025",
#     "Excess 2025",
#     "Ref. NO. 2025",
#     "Payment Method",  # duplicated
#     "Claim Reserves 2024",
#     "Voucher NO. 2024",
#     "Settled Amount 2024",
#     "Ref. 2024",
#     "Payment Method 2024",
#     "Excess 2024",
#     "Ref. NO. 2024",
#     "Payment Method",  # duplicated
#     "Claim Reserves 2023",
#     "Voucher NO. 2023",
#     "Settled Amount 2023",
#     "Ref. 2023",
#     "Payment Method 2023",
#     "Excess 2023",
#     "Ref. NO. 2023",
#     "Payment Method",  # duplicated
#     "Claim Reserves 2022",
#     "Voucher NO. 2022",
#     "Settled Amount 2022",
#     "Ref. 2022",
#     "Excess 2022",
#     "Ref. NO. 2022",
#     "Claim Reserves 2021",
#     "Voucher NO. 2021",
#     "Settled Amount 2021",
#     "Ref. 2021",
#     "Excess 2021",
#     "Ref. NO. 2021",
#     "Claim Reserves 2020",
#     "Voucher NO. 2020",
#     "Settled Amount 2020",
#     "Ref. 2020",
#     "Excess 2020",
#     "Ref. NO. 2020",
#     "Claim Reserves 2019",
#     "Voucher NO. 2019",
#     "Settled Amount 2019",
#     "Ref. 2019",
#     "Excess 2019",
#     "Ref. NO. 2019",
#     "Claim Reserves 2018",
#     "Voucher NO. 2018",
#     "Settled Amount 2018",
#     "Ref. 2018",
#     "Ref. NO. 2018",
#     "Claim Reserves 2017",
#     "Voucher NO. 2017",
#     "Settled Amount 2017",
#     "Ref. 2017",
#     "Excess 2017",
#     "Ref. NO. 2017",
#     "Voucher NO. 2016",
#     "Claim Reserves 2016",
#     "Settled Amount 2016",
#     "Ref. 2016",
#     "Excess 2016",
#     "Ref. NO. 2016",
#     "Voucher NO. 2015",
#     "Claim Reserves Dec 2015",
#     "Settled Amount Oct - Dec 2015",
#     "Ref. Oct - Dec 2015",
#     "Excess Oct - Dec 2015",
#     "Ref. NO. Oct - Dec 2015",
#     "Claim Reserves 2015 Sept",
#     "Voucher NO. 2015",
#     "Settled Amount 2015 (Jan-Sept)",
#     "Ref. NO. 2015",
#     "Excess 2015 (Jan-Sept)",
#     "Claim Reserves 2014",
#     "Voucher NO. 2014 (Reserve)",
#     "Payment 2014",
#     "Voucher NO. 2014 (Payment)",
#     "Excess 2014",
#     "Claim Reserves 2013",
#     "Voucher NO. 2013 (Reserve)",
#     "Settled Amount 2013",
#     "Voucher NO. 2013 (Payment)",
#     "Excess 2013",
#     "Claim Reserves 2012",
#     "Voucher NO. 2012 (Reserve)",
#     "Payment 2012",
#     "Voucher NO. 2012 (Payment)",
#     "Excess 2012",
#     "Claim Reserves 2011",
#     "Voucher NO. 2011 (Reserve)",
#     "Payment 2011",
#     "Voucher NO. 2011 (Payment)",
#     "Excess 2011",
#     "Claim Reserves 2010",
#     "Voucher NO. 2010 (Reserve)",
#     "Settled Amount 2010",
#     "Voucher NO. 2010 (Payment)",
#     "Excess 2010",
#     # Reinsurer columns
#     "",
#     "Reinsurer",
#     "Rein. Part. In Claims Reserves 2026",
#     "Voucher NO. 2026",
#     "RI Settled Amount 2026",
#     "Ref. 2026",
#     "Rein. Part. In Claims Reserves 2025",
#     "Voucher NO. 2025",
#     "RI Settled Amount 2025",
#     "Ref. 2025",
#     "Rein. Part. In Claims Reserves 2024",
#     "Voucher NO. 2024",
#     "RI Settled Amount 2024",
#     "Ref. 2024",
#     "Rein. Part. In Claims Reserves 2023",
#     "Voucher NO. 2023",
#     "RI Settled Amount 2023",
#     "Ref. 2023",
#     "Rein. Part. In Claims Reserves 2022",
#     "Voucher NO. 2022",
#     "RI Settled Amount 2022",
#     "Ref. 2022",
#     "Rein. Part. In Claims Reserves 2021",
#     "Voucher NO. 2021",
#     "RI Settled Amount 2021",
#     "Ref. 2021",
#     "Rein. Part. In Claims Reserves 2020",
#     "Voucher NO. 2020",
#     "RI Settled Amount 2020",
#     "Ref. 2020",
#     "Rein. Part. In Claims Reserves 2019",
#     "Voucher NO. 2019",
#     "RI Settled Amount 2019",
#     "Ref. 2019",
#     "Rein. Part. In Claims Reserves 2018",
#     "Voucher NO. 2018",
#     "RI Settled Amount 2018",
#     "Ref. 2018",
#     "Rein. Part. In Claims Reserves 2017",
#     "Voucher NO. 2017",
#     "RI Settled Amount 2017",
#     "Ref. 2017",
#     "Voucher NO. 2016",
#     "RI Settled Amount 2016",
#     "Ref. 2016",
#     "Voucher NO. 2015",
#     "RI Settled Amount Oct - Dec 2015",
#     "Ref. Oct - Dec 2015",
#     "RI Settled Amount 2015",
#     "Ref. 2015",
#     "Voucher NO. 2014",
#     "RI Settled Amount 2014",
#     "Ref. 2014",
#     "Voucher NO. 2013",
#     "RI Settled Amount 2013",
#     "Ref. 2013",
#     "Voucher NO. 2012",
#     "Voucher NO. 2011",
#     "Voucher NO. (before 2011)",
#     "RI Settled Amount 2011",
#     "Ref. 2011",
#     "RI Settled Amount 2012",
#     "Ref. 2012",
#     "Settled Amount (before 2011)",
#     "Ref. (before 2011)",
#     # Final summary columns
#     "Salvage", 
#     "Type",
#     "YearOc",
#     "YearRp",
#     "END_AC",
#     "Reserves",
#     "Payments",
#     "Excess",
#     "Salvage",
#     "Ceded Reserves",
#     "Ceded Payments",
#     "",
#     "Acc Res",
#     "Total paid",
#     "Acc excess",
#     "Ceded Res",
#     "Ceded pay-Excess",
#     "",
#     "retention pay",
#     "",
#     "DIFF",
#     "", "", 
#     "JV03-091",
#     "#N/A",
#     "", 
#     "/JV03-091",
#     "", "", 
#     "xol"
# ]

direct_data = pd.read_sql_query("SELECT * FROM direct_data", sqlite3.connect(r"C:\Users\F170\Desktop\AA Report\Function_Claims_Triangle\output\direct_data.db"))
fac_in_data = pd.read_sql_query("SELECT * FROM fac_in_data", sqlite3.connect(r"C:\Users\F170\Desktop\AA Report\Function_Claims_Triangle\output\fac_in_data.db"))
direct_data = direct_data.reindex(columns=TEMPLATE_COLUMNS_DATA_RECORD)
fac_in_data = fac_in_data.reindex(columns=TEMPLATE_COLUMNS_DATA_RECORD)

# Main Function - Triangle Generator
def triangle_generator(option, extract, show, product):
    original_direct_data = pd.read_sql_query("SELECT * FROM direct_data", sqlite3.connect(r"C:\Users\F170\Desktop\AA Report\Function_Claims_Triangle\output\direct_data.db"))
    original_fac_in_data = pd.read_sql_query("SELECT * FROM fac_in_data", sqlite3.connect(r"C:\Users\F170\Desktop\AA Report\Function_Claims_Triangle\output\fac_in_data.db"))
    
    original_direct_data['Type Sap'] = pd.to_numeric(original_direct_data['Type Sap'])
    original_fac_in_data['Type Sap'] = pd.to_numeric(original_fac_in_data['Type Sap'])
    original_direct_data['Date of Occurrence'] = pd.to_datetime(original_direct_data['Date of Occurrence'], errors='coerce', format = 'mixed')
    original_fac_in_data['Date of Occurrence'] = pd.to_datetime(original_fac_in_data['Date of Occurrence'], errors='coerce', format = 'mixed')
    original_direct_data['Accident yr'] = original_direct_data['Date of Occurrence'].dt.year
    original_fac_in_data['Accident yr'] = original_fac_in_data['Date of Occurrence'].dt.year

    amount_keywords = [
        "Settled Amount",
        "Excess",
        "RI Settled Amount",
        "Gr Clm Incurred",
        "RI Clm Incurred",
        "Claim Reserves",
    ]

    def force_numeric(df):
        cols = [
            c for c in df.columns
            if any(k in c for k in amount_keywords)
        ]

        for c in cols:
            df[c] = (
                df[c]
                .astype(str)
                .str.replace(",", "", regex=True)
                .replace({"": None, " ": None})
                .pipe(pd.to_numeric, errors="coerce")
            )

        return df
    
    original_direct_data = force_numeric(original_direct_data)
    original_fac_in_data = force_numeric(original_fac_in_data)

    gross_payment_matrix                        = pd.DataFrame(index=years, columns=years, dtype=float)
    ri_payment_matrix                           = pd.DataFrame(index=years, columns=years, dtype=float)
    net_payment_matrix                          = pd.DataFrame(index=years, columns=years, dtype=float)
    gross_incurred_matrix                       = pd.DataFrame(index=years, columns=years, dtype=float)
    ri_incurred_matrix                          = pd.DataFrame(index=years, columns=years, dtype=float)
    net_incurred_matrix                         = pd.DataFrame(index=years, columns=years, dtype=float)
    direct_gross_payment_matrix                 = pd.DataFrame(index=years, columns=years, dtype=float)
    direct_ri_payment_matrix                    = pd.DataFrame(index=years, columns=years, dtype=float)
    direct_net_payment_matrix                   = pd.DataFrame(index=years, columns=years, dtype=float)
    direct_gross_incurred_matrix                = pd.DataFrame(index=years, columns=years, dtype=float)
    direct_ri_incurred_matrix                   = pd.DataFrame(index=years, columns=years, dtype=float)
    direct_net_incurred_matrix                  = pd.DataFrame(index=years, columns=years, dtype=float)
    fac_in_gross_payment_matrix                 = pd.DataFrame(index=years, columns=years, dtype=float)
    fac_in_ri_payment_matrix                    = pd.DataFrame(index=years, columns=years, dtype=float)
    fac_in_net_payment_matrix                   = pd.DataFrame(index=years, columns=years, dtype=float)
    fac_in_gross_incurred_matrix                = pd.DataFrame(index=years, columns=years, dtype=float)
    fac_in_ri_incurred_matrix                   = pd.DataFrame(index=years, columns=years, dtype=float)
    fac_in_net_incurred_matrix                  = pd.DataFrame(index=years, columns=years, dtype=float)
    reported_case_count_matrix                  = pd.DataFrame(index=years, columns=years, dtype=float)
    closed_claims_count_matrix                  = pd.DataFrame(index=years, columns=years, dtype=float)
    direct_outstanding_claims_count_matrix      = pd.DataFrame(index=years, columns=years, dtype=float)
    fac_in_outstanding_claims_count_matrix      = pd.DataFrame(index=years, columns=years, dtype=float)

    gross_payment_cum_matrix                    = pd.DataFrame(index=years, columns=years, dtype=float)
    ri_payment_cum_matrix                       = pd.DataFrame(index=years, columns=years, dtype=float)
    net_payment_cum_matrix                      = pd.DataFrame(index=years, columns=years, dtype=float)
    gross_incurred_cum_matrix                   = pd.DataFrame(index=years, columns=years, dtype=float)
    ri_incurred_cum_matrix                      = pd.DataFrame(index=years, columns=years, dtype=float)
    net_incurred_cum_matrix                     = pd.DataFrame(index=years, columns=years, dtype=float)
    direct_gross_payment_cum_matrix             = pd.DataFrame(index=years, columns=years, dtype=float)
    direct_ri_payment_cum_matrix                = pd.DataFrame(index=years, columns=years, dtype=float)
    direct_net_payment_cum_matrix               = pd.DataFrame(index=years, columns=years, dtype=float)
    direct_gross_incurred_cum_matrix            = pd.DataFrame(index=years, columns=years, dtype=float)
    direct_ri_incurred_cum_matrix               = pd.DataFrame(index=years, columns=years, dtype=float)
    direct_net_incurred_cum_matrix              = pd.DataFrame(index=years, columns=years, dtype=float)
    fac_in_gross_payment_cum_matrix             = pd.DataFrame(index=years, columns=years, dtype=float)
    fac_in_ri_payment_cum_matrix                = pd.DataFrame(index=years, columns=years, dtype=float)
    fac_in_net_payment_cum_matrix               = pd.DataFrame(index=years, columns=years, dtype=float)
    fac_in_gross_incurred_cum_matrix            = pd.DataFrame(index=years, columns=years, dtype=float)
    fac_in_ri_incurred_cum_matrix               = pd.DataFrame(index=years, columns=years, dtype=float)
    fac_in_net_incurred_cum_matrix              = pd.DataFrame(index=years, columns=years, dtype=float)
    reported_case_count_cum_matrix              = pd.DataFrame(index=years, columns=years, dtype=float)
    closed_claims_count_cum_matrix              = pd.DataFrame(index=years, columns=years, dtype=float)

    if option == 'All':
        # Account Year Triangle
        for i in range(len(years)):

            year = years[i]
            direct_origin = original_direct_data[
                original_direct_data['Accident yr'] == year
            ]

            fac_in_origin = original_fac_in_data[
                original_fac_in_data['Accident yr'] == year
            ]   

            for j in range(len(years)):
                calendar_year = years[j]
            
                settle_col = f"Settled Amount \n{calendar_year}"
                excess_col = f"Excess \n{calendar_year}"
                ri_settled_col = f"RI Settled Amount \n{calendar_year}"
                incurr_col = f"Gr Clm Incurred \n{calendar_year}"
                ri_incurr_col = f"RI Clm Incurred \n{calendar_year}"

                gross_payment_matrix.loc[year, calendar_year] = (
                    direct_origin[settle_col].sum() +
                    direct_origin[excess_col].sum() +
                    fac_in_origin[settle_col].sum() + 
                    fac_in_origin[excess_col].sum()
                )

                ri_payment_matrix.loc[year, calendar_year] = (
                    direct_origin[ri_settled_col].sum() +
                    fac_in_origin[ri_settled_col].sum()
                )

                gross_incurred_matrix.loc[year, calendar_year] = (
                    direct_origin[incurr_col].sum() +
                    fac_in_origin[incurr_col].sum()
                )

                ri_incurred_matrix.loc[year, calendar_year] = (
                    direct_origin[ri_incurr_col].sum() +
                    fac_in_origin[ri_incurr_col].sum()
                )

                net_payment_matrix = gross_payment_matrix - ri_payment_matrix
                net_incurred_matrix = gross_incurred_matrix - ri_incurred_matrix

        # Development Year Triangle
        for l in range(len(years)):
            year = years[l]
            direct_origin = original_direct_data[
                original_direct_data['Accident yr'] == year
            ]

            fac_in_origin = original_fac_in_data[
                original_fac_in_data['Accident yr'] == year
            ]   

            for m in range(len(years)):
                start_year = years[l]
                end_year   = years[m]

                settle_col = [
                    f"Settled Amount \n{y}" 
                    for y in range(start_year, end_year + 1)
                    if f"Settled Amount \n{y}" in direct_origin.columns
                ]
                excess_col = [
                    f"Excess \n{y}"
                    for y in range(start_year, end_year + 1)
                    if f"Excess \n{y}" in direct_origin.columns
                ]
                ri_col = [
                    f"RI Settled Amount \n{y}"
                    for y in range(start_year, end_year + 1)
                    if f"RI Settled Amount \n{y}" in direct_origin.columns
                ]

                incurr_col = [
                    f"Gr Clm Incurred \n{y}"
                    for y in range(start_year, end_year + 1)
                    if f"Gr Clm Incurred \n{y}" in direct_origin.columns
                ]

                ri_incurr_col = [
                    f"RI Clm Incurred \n{y}"
                    for y in range(start_year, end_year + 1)
                    if f"RI Clm Incurred \n{y}" in direct_origin.columns
                ]
                # compute the total once
                total_val = (
                    direct_origin[settle_col].sum().sum() +
                    direct_origin[excess_col].sum().sum() +
                    fac_in_origin[settle_col].sum().sum() +
                    fac_in_origin[excess_col].sum().sum()
                )

                ri_val = (
                    direct_origin[ri_col].sum().sum() + 
                    fac_in_origin[ri_col].sum().sum()
                )

                incurr_val = (
                    direct_origin[incurr_col].sum().sum() +
                    fac_in_origin[incurr_col].sum().sum()
                )
                ri_incurr_val = (
                    direct_origin[ri_incurr_col].sum().sum() +
                    fac_in_origin[ri_incurr_col].sum().sum()
                )
                # shift each row (except first) left by its row index
                col_idx = m - l
                if 0 <= col_idx < len(years):
                    gross_payment_cum_matrix.iloc[l, col_idx] = total_val
                    ri_payment_cum_matrix.iloc[l, col_idx] = ri_val
                    gross_incurred_cum_matrix.iloc[l, col_idx] = incurr_val
                    ri_incurred_cum_matrix.iloc[l, col_idx] = ri_incurr_val
                net_payment_cum_matrix = gross_payment_cum_matrix - ri_payment_cum_matrix
                net_incurred_cum_matrix = gross_incurred_cum_matrix - ri_incurred_cum_matrix

        triangles = {
            'all_lines_gross_payment_account_year_triangle': gross_payment_matrix,
            'all_lines_ri_payment_account_year_triangle': ri_payment_matrix,
            'all_lines_net_payment_account_year_triangle': net_payment_matrix,
            'all_lines_gross_payment_development_year_triangle': gross_payment_cum_matrix,
            'all_lines_ri_payment_development_year_triangle': ri_payment_cum_matrix,
            'all_lines_net_payment_development_year_triangle': net_payment_cum_matrix,

            'all_lines_gross_incurred_account_year_triangle': gross_incurred_matrix,
            'all_lines_ri_incurred_account_year_triangle': ri_incurred_matrix,
            'all_lines_net_incurred_account_year_triangle': net_incurred_matrix,
            'all_lines_gross_incurred_development_year_triangle': gross_incurred_cum_matrix,
            'all_lines_ri_incurred_development_year_triangle': ri_incurred_cum_matrix,
            'all_lines_net_incurred_development_year_triangle': net_incurred_cum_matrix,
        }

        with sqlite3.connect(r"C:\Users\F170\Desktop\AA Report\DBs\all_triangles.db") as conn:
            for table_name, tri in triangles.items():
                tri.to_sql(
                    name = table_name,
                    con = conn,
                    if_exists = 'replace',
                    index = False,
                    method = 'multi'
                )
                
    elif option == 'Single':
        for i, (sap_code_str, product_name) in enumerate(list_of_lob.items()):
            if extract:
                original_direct_data = original_direct_data[original_direct_data['Type Sap'] == 43].copy()
                original_fac_in_data = original_fac_in_data[original_fac_in_data['Type Sap'] == 43].copy()

                if not show:
                    original_direct_data_single = original_direct_data[~original_direct_data['Product'].isin(['MACI','MPCC'])].copy()
                    original_fac_in_data_single = original_fac_in_data[~original_fac_in_data['Product'].isin(['MACI','MPCC'])].copy()
                elif product in ['MACI','MPCC']:
                    original_direct_data_single = original_direct_data[original_direct_data['Product'].isin([product])].copy()
                    original_fac_in_data_single = original_fac_in_data[original_fac_in_data['Product'].isin([product])].copy()
                else:
                    raise ValueError(f"Invalid product '{product}' when extract=True and show=True. Must be 'MACI' or 'MPCC'.")

            else:
                sap_code_str = pd.to_numeric(sap_code_str)
                original_direct_data_single = original_direct_data[original_direct_data['Type Sap'] == sap_code_str].copy()
                original_fac_in_data_single = original_fac_in_data[original_fac_in_data['Type Sap'] == sap_code_str].copy()
        
            # Account Year Triangle
            for p in range(len(years)):
                year = years[p]
                direct_origin = original_direct_data_single[
                    original_direct_data_single['Accident yr'] == year
                ]

                fac_in_origin = original_fac_in_data_single[
                    original_fac_in_data_single['Accident yr'] == year
                ]
                
                for j in range(len(years)):
                    calendar_year = years[j]
                    
                    settle_col     = f"Settled Amount \n{calendar_year}"
                    excess_col     = f"Excess \n{calendar_year}"
                    ri_settled_col = f"RI Settled Amount \n{calendar_year}"
                    incurr_col     = f"Gr Clm Incurred \n{calendar_year}"
                    ri_incurr_col  = f"RI Clm Incurred \n{calendar_year}"

                    gross_payment_matrix.loc[year, calendar_year] = (
                        direct_origin[settle_col].sum().sum() +
                        direct_origin[excess_col].sum().sum() +
                        fac_in_origin[settle_col].sum().sum() + 
                        fac_in_origin[excess_col].sum()
                    )

                    ri_payment_matrix.loc[year, calendar_year] = (
                        direct_origin[ri_settled_col].sum() +
                        fac_in_origin[ri_settled_col].sum()
                    )

                    direct_gross_payment_matrix.loc[year, calendar_year] = (
                        direct_origin[settle_col].sum() +
                        direct_origin[excess_col].sum()
                    )

                    direct_ri_payment_matrix.loc[year, calendar_year] = (
                        direct_origin[ri_settled_col].sum()
                    )

                    fac_in_gross_payment_matrix.loc[year, calendar_year] = (
                        fac_in_origin[settle_col].sum()
                    )

                    fac_in_ri_payment_matrix.loc[year, calendar_year] = (
                        fac_in_origin[ri_settled_col].sum()
                    )
                    
                    ri_incurred_matrix.loc[year, calendar_year] = (
                        direct_origin[ri_incurr_col].sum() +
                        fac_in_origin[ri_incurr_col].sum()
                    )

                    direct_ri_incurred_matrix.loc[year, calendar_year] = (
                        direct_origin[ri_settled_col].sum()
                    )

                    if year == 2010:
                        direct_origin[f'Claim Reserves \n{year - 1}'] = 0
                        fac_in_origin[f'Claim Reserves \n{year - 1}'] = 0
                        
                    direct_gross_incurred_matrix.loc[year, calendar_year] = (
                        direct_origin[f'Claim Reserves \n{year}'].sum() + 
                        direct_origin[f'Claim Reserves \n{year - 1}'].sum() + 
                        direct_origin[f'Settled Amount \n{year}'].sum() + 
                        direct_origin[f'Excess \n{year}'].sum()
                    )

                    fac_in_gross_incurred_matrix.loc[year, calendar_year] = (
                        fac_in_origin[f'Claim Reserves \n{year}'].sum() + 
                        fac_in_origin[f'Claim Reserves \n{year - 1}'].sum() + 
                        fac_in_origin[f'Settled Amount \n{year}'].sum() + 
                        fac_in_origin[f'Excess \n{year}'].sum()
                    )
                    
                    gross_incurred_matrix.loc[year, calendar_year] = (
                        direct_gross_incurred_matrix +
                        fac_in_gross_incurred_matrix
                    )
                    fac_in_ri_incurred_matrix.loc[year, calendar_year] = (
                        fac_in_origin[ri_settled_col].sum()
                    )

                    reported_case_count_matrix.loc[year, calendar_year] = (
                        ((direct_origin['Op Year'] == calendar_year) & 
                         (direct_origin['Accident yr'] == year) & 
                         (direct_origin['Valid case #'] == 1)).sum() + 
                        ((fac_in_origin['Op Year'] == calendar_year) & 
                         (fac_in_origin['Accident yr'] == year) & 
                         (fac_in_origin['Valid case #'] == 1)).sum()
                    )

                    closed_claims_count_matrix.loc[year, calendar_year] = (
                        ((direct_origin[f'Claim Reserves \n{calendar_year}'] == 0) & 
                         (direct_origin['Accident yr'] == year) & 
                         (direct_origin['Valid case #'] == 1)).sum() + 
                        ((fac_in_origin[f'Claim Reserves \n{calendar_year}'] == 0) & 
                         (fac_in_origin['Accident yr'] == year) & 
                         (fac_in_origin['Valid case #'] == 1)).sum()
                    )

                    direct_outstanding_claims_count_matrix.loc[year, calendar_year] = (
                        ((direct_origin['Accident yr'] == year) & 
                         (direct_origin['Status'] == 'Open') &                          
                         (direct_origin['Valid case #'] == 1)).sum()
                    )

                    fac_in_outstanding_claims_count_matrix.loc[year, calendar_year] = (
                        ((fac_in_origin['Accident yr'] == year) & 
                         (fac_in_origin['Status'] == 'Open') &                          
                         (fac_in_origin['Valid case #'] == 1)).sum()
                    )
                    
                    net_payment_matrix = gross_payment_matrix - ri_payment_matrix
                    direct_net_payment_matrix = direct_gross_payment_matrix - direct_ri_payment_matrix
                    fac_in_net_payment_matrix = fac_in_gross_payment_matrix - fac_in_ri_payment_matrix

                    net_incurred_matrix = gross_incurred_matrix - ri_incurred_matrix
                    direct_net_incurred_matrix = direct_gross_incurred_matrix - direct_ri_incurred_matrix
                    fac_in_net_incurred_matrix = fac_in_gross_incurred_matrix - fac_in_ri_incurred_matrix
                    
            # Development Year Triangle
            for l in range(len(years)):
                year = years[l]
                direct_origin = original_direct_data_single[
                    original_direct_data_single['Accident yr'] == year
                ]

                fac_in_origin = original_fac_in_data_single[
                    original_fac_in_data_single['Accident yr'] == year
                ]

                for m in range(len(years)):
                    start_year = years[l]
                    end_year   = years[m]

                    settle_col = [
                        f"Settled Amount \n{y}" 
                        for y in range(start_year, end_year + 1)
                        if f"Settled Amount \n{y}" in direct_origin.columns
                    ]
                    excess_col = [
                        f"Excess \n{y}"
                        for y in range(start_year, end_year + 1)
                        if f"Excess \n{y}" in direct_origin.columns
                    ]
                    ri_col = [
                        f"RI Settled Amount \n{y}"
                        for y in range(start_year, end_year + 1)
                        if f"RI Settled Amount \n{y}" in direct_origin.columns
                    ]
                    incurr_col = [
                        f"Gr Clm Incurred \n{y}"
                        for y in range(start_year, end_year + 1)
                        if f"Gr Clm Incurred \n{y}" in direct_origin.columns
                    ]

                    ri_incurr_col = [
                        f"RI Clm Incurred \n{y}"
                        for y in range(start_year, end_year + 1)
                        if f"RI Clm Incurred \n{y}" in direct_origin.columns
                    ]
                    # compute the total once
                    total_val = (
                        direct_origin[settle_col].sum().sum() +
                        direct_origin[excess_col].sum().sum() +
                        fac_in_origin[settle_col].sum().sum() +
                        fac_in_origin[excess_col].sum().sum()
                    )

                    ri_val = (
                        direct_origin[ri_col].sum().sum() + 
                        fac_in_origin[ri_col].sum().sum()
                    )

                    ri_incurr_val = (
                        direct_origin[ri_incurr_col].sum().sum() +
                        fac_in_origin[ri_incurr_col].sum().sum()
                    )

                    direct_total_val = (
                        direct_origin[settle_col].sum().sum() +
                        direct_origin[excess_col].sum().sum() 
                    )

                    direct_ri_total_val = (
                        direct_origin[ri_col].sum().sum()
                    )

                    if year == 2010:
                        direct_origin[f'Claim Reserves \n{year - 1}'] = 0
                        fac_in_origin[f'Claim Reserves \n{year - 1}'] = 0
                        
                    direct_incurr_val = (
                        direct_origin[f'Claim Reserves \n{year}'].sum().sum() + 
                        direct_origin[f'Claim Reserves \n{year - 1}'].sum().sum() + 
                        direct_origin[f'Settled Amount \n{year}'].sum().sum() + 
                        direct_origin[f'Excess \n{year}'].sum().sum()
                    )

                    direct_ri_incurr_val = (
                        direct_origin[ri_incurr_col].sum().sum()
                    )

                    fac_in_total_val = (
                        fac_in_origin[settle_col].sum().sum()
                    )

                    fac_in_incurr_val = (
                        fac_in_origin[f'Claim Reserves \n{year}'].sum().sum() + 
                        fac_in_origin[f'Claim Reserves \n{year - 1}'].sum().sum() + 
                        fac_in_origin[f'Settled Amount \n{year}'].sum().sum() + 
                        fac_in_origin[f'Excess \n{year}'].sum().sum()
                    )

                    incurr_val = (
                        direct_incurr_val +
                        fac_in_incurr_val
                    )


                    fac_in_ri_incurr_val = (
                        fac_in_origin[ri_incurr_col].sum().sum()
                    )

                    fac_in_ri_total_val = (
                        fac_in_origin[ri_col].sum().sum()
                    )

                    reported_case_count_val = (
                        ((direct_origin['Op Year'] <= end_year) & 
                         (direct_origin['Accident yr'] == start_year) & 
                         (direct_origin['Valid case #'] == 1)).sum().sum() + 
                        ((fac_in_origin['Op Year'] <= end_year) & 
                         (fac_in_origin['Accident yr'] == start_year) & 
                         (fac_in_origin['Valid case #'] == 1)).sum().sum()
                    )

                    closed_claims_count_val = (
                        ((direct_origin['Accident yr'] == start_year) & 
                         (direct_origin[f'Claim Reserves \n{end_year}'] == 0) & 
                         (direct_origin['Valid case #'] == 1)).sum().sum() + 
                        ((fac_in_origin['Accident yr'] == start_year) & 
                         (fac_in_origin[f'Claim Reserves \n{end_year}'] == 0) &
                         (fac_in_origin['Valid case #'] == 1)).sum().sum()
                    )


                    # shift each row (except first) left by its row index
                    col_idx = m - l 
                    if 0 <= col_idx < len(years):
                        gross_payment_cum_matrix.iloc[l, col_idx] = total_val
                        ri_payment_cum_matrix.iloc[l, col_idx] = ri_val

                        direct_gross_payment_cum_matrix.iloc[l, col_idx] = direct_total_val
                        direct_ri_payment_cum_matrix.iloc[l, col_idx] = direct_ri_total_val

                        fac_in_gross_payment_cum_matrix.iloc[l, col_idx] = fac_in_total_val
                        fac_in_ri_payment_cum_matrix.iloc[l, col_idx] = fac_in_ri_total_val

                        gross_incurred_cum_matrix.iloc[l, col_idx] = incurr_val
                        ri_incurred_cum_matrix.iloc[l, col_idx] = ri_incurr_val

                        direct_gross_incurred_cum_matrix.iloc[l, col_idx] = direct_incurr_val
                        direct_ri_incurred_cum_matrix.iloc[l, col_idx] = direct_ri_incurr_val

                        fac_in_gross_incurred_cum_matrix.iloc[l, col_idx] = fac_in_incurr_val
                        fac_in_ri_incurred_cum_matrix.iloc[l, col_idx] = fac_in_ri_incurr_val

                        reported_case_count_cum_matrix.iloc[l, col_idx] = reported_case_count_val
                        closed_claims_count_cum_matrix.iloc[l, col_idx] = closed_claims_count_val

                    net_payment_cum_matrix = gross_payment_cum_matrix - ri_payment_cum_matrix
                    direct_net_payment_cum_matrix = direct_gross_payment_cum_matrix - direct_ri_payment_cum_matrix
                    fac_in_net_payment_cum_matrix = fac_in_gross_payment_cum_matrix - fac_in_ri_payment_cum_matrix

                    net_incurred_cum_matrix = gross_incurred_cum_matrix - ri_incurred_cum_matrix
                    direct_net_incurred_cum_matrix = direct_gross_incurred_cum_matrix - direct_ri_incurred_cum_matrix
                    fac_in_net_incurred_cum_matrix = fac_in_gross_incurred_cum_matrix - fac_in_ri_incurred_cum_matrix

            triangles = {
                f'{sap_code_str}_{product_name}_gross_payment_account_year_triangle': gross_payment_matrix,
                f'{sap_code_str}_{product_name}_ri_payment_account_year_triangle': ri_payment_matrix,
                f'{sap_code_str}_{product_name}_net_payment_account_year_triangle': net_payment_matrix,
                f'{sap_code_str}_{product_name}_gross_payment_development_year_triangle': gross_payment_cum_matrix,
                f'{sap_code_str}_{product_name}_ri_payment_development_year_triangle': ri_payment_cum_matrix,
                f'{sap_code_str}_{product_name}_net_payment_development_year_triangle': net_payment_cum_matrix,

                f'{sap_code_str}_{product_name}_direct_gross_payment_account_year_triangle': direct_gross_payment_matrix,
                f'{sap_code_str}_{product_name}_direct_ri_payment_account_year_triangle': direct_ri_payment_matrix,
                f'{sap_code_str}_{product_name}_direct_net_payment_account_year_triangle': direct_net_payment_matrix,
                f'{sap_code_str}_{product_name}_direct_gross_payment_development_year_triangle': direct_gross_payment_cum_matrix,
                f'{sap_code_str}_{product_name}_direct_ri_payment_development_year_triangle': direct_ri_payment_cum_matrix,
                f'{sap_code_str}_{product_name}_direct_net_payment_development_year_triangle': direct_net_payment_cum_matrix,

                f'{sap_code_str}_{product_name}_fac_in_gross_payment_account_year_triangle': fac_in_gross_payment_matrix,
                f'{sap_code_str}_{product_name}_fac_in_ri_payment_account_year_triangle': fac_in_ri_payment_matrix,
                f'{sap_code_str}_{product_name}_fac_in_net_payment_account_year_triangle': fac_in_net_payment_matrix,
                f'{sap_code_str}_{product_name}_fac_in_gross_payment_development_year_triangle': fac_in_gross_payment_cum_matrix,
                f'{sap_code_str}_{product_name}_fac_in_ri_payment_development_year_triangle': fac_in_ri_payment_cum_matrix,
                f'{sap_code_str}_{product_name}_fac_in_net_payment_development_year_triangle': fac_in_net_payment_cum_matrix,

                f'{sap_code_str}_{product_name}_gross_incurred_account_year_triangle': gross_incurred_matrix,
                f'{sap_code_str}_{product_name}_ri_incurred_account_year_triangle': ri_incurred_matrix,
                f'{sap_code_str}_{product_name}_net_incurred_account_year_triangle': net_incurred_matrix,
                f'{sap_code_str}_{product_name}_gross_incurred_development_year_triangle': gross_incurred_cum_matrix,
                f'{sap_code_str}_{product_name}_ri_incurred_development_year_triangle': ri_incurred_cum_matrix,
                f'{sap_code_str}_{product_name}_net_incurred_development_year_triangle': net_incurred_cum_matrix,

                f'{sap_code_str}_{product_name}_direct_gross_incurred_account_year_triangle': direct_gross_incurred_matrix,
                f'{sap_code_str}_{product_name}_direct_ri_incurred_account_year_triangle': direct_ri_incurred_matrix,
                f'{sap_code_str}_{product_name}_direct_net_incurred_account_year_triangle': direct_net_incurred_matrix,
                f'{sap_code_str}_{product_name}_direct_gross_incurred_development_year_triangle': direct_gross_incurred_cum_matrix,
                f'{sap_code_str}_{product_name}_direct_ri_incurred_development_year_triangle': direct_ri_incurred_cum_matrix,
                f'{sap_code_str}_{product_name}_direct_net_incurred_development_year_triangle': direct_net_incurred_cum_matrix,

                f'{sap_code_str}_{product_name}_fac_in_gross_incurred_account_year_triangle': fac_in_gross_incurred_matrix,
                f'{sap_code_str}_{product_name}_fac_in_ri_incurred_account_year_triangle': fac_in_ri_incurred_matrix,
                f'{sap_code_str}_{product_name}_fac_in_net_incurred_account_year_triangle': fac_in_net_incurred_matrix,
                f'{sap_code_str}_{product_name}_fac_in_gross_incurred_development_year_triangle': fac_in_gross_incurred_cum_matrix,
                f'{sap_code_str}_{product_name}_fac_in_ri_incurred_development_year_triangle': fac_in_ri_incurred_cum_matrix,
                f'{sap_code_str}_{product_name}_fac_in_net_incurred_development_year_triangle': fac_in_net_incurred_cum_matrix,

                f'{sap_code_str}_{product_name}_case_count_account_year_triangle': reported_case_count_matrix,
                f'{sap_code_str}_{product_name}_case_count_development_year_triangle': reported_case_count_cum_matrix,
                f'{sap_code_str}_{product_name}_closed_claims_count_account_year_triangle': closed_claims_count_matrix,
                f'{sap_code_str}_{product_name}_closed_claims_count_development_year_triangle': closed_claims_count_cum_matrix,
                f'{sap_code_str}_{product_name}_direct_outstanding_claims_count_year_triangle': direct_outstanding_claims_count_matrix,
                f'{sap_code_str}_{product_name}_fac_in_outstanding_claims_count_year_triangle': fac_in_outstanding_claims_count_matrix
               
            }
            if extract == True:
                if show == False:
                    with sqlite3.connect(r"C:\Users\F170\Desktop\AA Report\Function_Claims_Triangle\output\no_maci_and_mpcc.db") as conn:
                        for table_name, tri in triangles.items():
                            tri.to_sql(
                                name = table_name,
                                con = conn,
                                if_exists = 'replace',
                                index = False,
                                method = 'multi'
                            )
                elif show == True:
                    with sqlite3.connect(fr"C:\Users\F170\Desktop\AA Report\Function_Claims_Triangle\output\{product}.db") as conn:
                        for table_name, tri in triangles.items():
                            tri.to_sql(
                                name = table_name,
                                con = conn,
                                if_exists = 'replace',
                                index = False,
                                method = 'multi'
                            )
            else:
                with sqlite3.connect(r"C:\Users\F170\Desktop\AA Report\Function_Claims_Triangle\output\all_triangles.db") as conn:
                    for table_name, tri in triangles.items():
                        tri.to_sql(
                            name = table_name,
                            con = conn,
                            if_exists = 'replace',
                            index = False,
                            method = 'multi'
                    )

# Output report for EY
def report_ibnr_projection(option):
    # 1. Configuration
    db_path = "all_triangles.db" 
    template_path = "template_rbc.xlsm"
    output_name = f"output_{option.replace(' ', '_')}.xlsm"
    
    mapping = {
        'EC': ['2111'],
        'Income Protection': ['2121', '2122', '2123'],
        'Medical Expenses': ['222', '223'],
        'Fire': ['311', '323', '3261', '3262', '3263', '32711', '3273', '3283'],
        'Other Motor': ['41'],
        'Motor TPL': ['43', '44'],
        'Transport': ['62', '71', '72', '79'],
        'General Liability': ['82', '83', '854', '92']
    }
    
    sap_codes = mapping.get(option, [])
    sums = {k: None for k in ['closed', 'reported', 'g_pay', 'g_inc', 'n_pay', 'n_inc']}

    # 2. SQL Data Processing
    conn = sqlite3.connect(db_path)
    for code in sap_codes:
        # Assuming list_of_lob is defined globally or imported
        product_name = list_of_lob[int(code)]
        
        tables = {
            'closed': f'"{code}_{product_name}_closed_claims_count_development_year_triangle"',
            'reported': f'"{code}_{product_name}_case_count_development_year_triangle"',
            'g_pay': f'"{code}_{product_name}_gross_payment_development_year_triangle"',
            'g_inc': f'"{code}_{product_name}_gross_incurred_development_year_triangle"',
            'n_pay': f'"{code}_{product_name}_net_payment_development_year_triangle"',
            'n_inc': f'"{code}_{product_name}_net_incurred_development_year_triangle"'
        }

        for key, table_name in tables.items():
            df = pd.read_sql_query(f'SELECT * FROM {table_name}', conn)
            # Cleanup logic
            n = df.shape[0]
            for i in range(n):
                df.iat[n-1-i, i] = 0
            df = df.iloc[:-1, :-1]

            if sums[key] is None: sums[key] = df
            else: sums[key] += df
    conn.close()

    # 3. Excel Injection
    wb = load_workbook(template_path, keep_vba=True)
    ws = wb['Data']
    ws.cell(row=2, column=2, value=option)

    for i in range(sums['g_pay'].shape[0]):
        for j in range(sums['g_pay'].shape[1]):
            ws.cell(row=43 + i,  column=5 + j, value=sums['g_pay'].iloc[i, j])
            ws.cell(row=95 + i,  column=5 + j, value=sums['g_inc'].iloc[i, j])
            ws.cell(row=149 + i, column=5 + j, value=sums['n_pay'].iloc[i, j])
            ws.cell(row=201 + i, column=5 + j, value=sums['n_inc'].iloc[i, j])
            ws.cell(row=254 + i, column=5 + j, value=sums['closed'].iloc[i, j])
            ws.cell(row=306 + i, column=5 + j, value=sums['reported'].iloc[i, j])

    wb.save(output_name)
    return output_name

# Output report for RBC
# Missing Outstanding Claims Count
# Missing data reconciliation
def report_claims_reserve():
    data_direct = pd.read_excel(r'Function_Claims_Triangle/output/claims_triangle.xlsx', 
                                sheet_name = 'Direct Data')
    data_fac_in = pd.read_excel(r'Function_Claims_Triangle/output/claims_triangle.xlsx', 
                                sheet_name = 'Fac-in Data')
    data_direct = data_direct[(data_direct['Type Sap'] == 43) | (data_direct['Type Sap'] == 2111)]
    data_fac_in = data_fac_in[(data_fac_in['Type Sap'] == 43) | (data_fac_in['Type Sap'] == 2111)] 
    template = r"C:\Users\F170\Desktop\AA Report\Function_Pillar I\template\claims_reserve_template.xlsx"
    output = r"C:\Users\F170\Desktop\AA Report\Function_Pillar I\output\claims_reserve_kai.xlsx"
    gpe_data = pd.read_excel(r"C:\Users\F170\Desktop\AA Report\Function_Pillar I\template\GPE 2026 (updated to March).xlsx",
                                sheet_name = 'GPE')
    gpe_sub = pd.read_excel(r"C:\Users\F170\Desktop\AA Report\Function_Pillar I\template\P&L 2026-03 NL (Get Closer).xlsx",
                            sheet_name = 'NL by line')
    
    gpe_direct_motor = gpe_data.iloc[5, 30]
    gpe_net_motor = gpe_sub.iloc[11, 18]
    gpe_mpcc = gpe_data.iloc[13, 30]
    gpe_maci = gpe_data.iloc[21, 30]
    gpe_direct_ec = gpe_data.iloc[29, 30]
    gpe_fac_in_ec = gpe_data.iloc[37, 30]

    wb = load_workbook(template)

    # Define your formula logic steps
    gr_clm_incurred_steps = [
            ('K', None, 'AB', 'AS'), ('L', 'K', 'AC', 'AT'), ('M', 'L', 'AD', 'AU'),
            ('N', 'M', 'AE', 'AV'), ('O', 'N', 'AF', 'AW'), ('P', 'O', 'AG', 'AX'),
            ('Q', 'P', 'AH', 'AY'), ('R', 'Q', 'AI', 'AZ'), ('S', 'R', 'AJ', 'BA'),
            ('T', 'S', 'AK', 'BB'), ('U', 'T', 'AL', 'BC'), ('V', 'U', 'AM', 'BD'),
            ('W', 'V', 'AN', 'BE'), ('X', 'W', 'AO', 'BF'), ('Y', 'X', 'AP', 'BG'),
            ('Z', 'Y', 'AQ', 'BH'), ('AA', 'Z', 'AR', 'BI')
    ]

    ri_clm_incurred_steps = [
        ('BJ', None, 'CA'), ('BK', 'BJ', 'CB'), ('BL', 'BK', 'CC'), ('BM', 'BL', 'CD'), ('BN', 'BM', 'CE'), 
        ('BO', 'BN', 'CF'), ('BP', 'BO', 'CG'), ('BQ', 'BP', 'CH'), ('BR', 'BQ', 'CI'), ('BS', 'BR', 'CJ'),
        ('BT', 'BS', 'CK'), ('BU', 'BT', 'CL'), ('BV', 'BU', 'CM'), ('BW', 'BV', 'CN'), ('BX', 'BW', 'CO'),
        ('BY', 'BX', 'CP'), ('BZ', 'BY', 'CQ')
    ]

    net_clm_incurred_steps = [
        ('DB', 'DT'), ('DC', 'DU'), ('DB', 'DT'), ('DD', 'DV'), ('DE', 'DW'), ('DF', 'DX'), ('DG', 'DY'), ('DH', 'DZ')
    ]

    # Map the sheets to their dataframes
    tasks_config = [
        ('Direct Data', data_direct),
        ('Fac-in Data', data_fac_in)
    ]

    for sheet_name, df in tasks_config:
        ws = wb[sheet_name]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                if isinstance(value, (int, float)) or value is None:
                    ws.cell(row=r_idx, column=c_idx).value = value
                    continue
                try:
                    ws.cell(row=r_idx, column=c_idx).value = float(value)
                except (ValueError, TypeError):
                    ws.cell(row=r_idx, column=c_idx).value = value
            
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            for i, (curr, prev, c3, c4) in enumerate(gr_clm_incurred_steps):
                gr_clm_target_col = 96 + i 
                if prev is None:
                    formula_gr_clm_incurred = f"=ROUND({curr}{r_idx}+{c3}{r_idx}+{c4}{r_idx},2)"
                else:
                    formula_gr_clm_incurred = f"=ROUND({curr}{r_idx}-{prev}{r_idx}+{c3}{r_idx}+{c4}{r_idx},2)"
                
                ws.cell(row=r_idx, column=gr_clm_target_col).value = formula_gr_clm_incurred

            for i, (curr, prev, c3) in enumerate(ri_clm_incurred_steps):
                ri_clm_target_col = 114 + i
                if prev is None:
                    formula_ri_clm_incurred = f"=ROUND({curr}{r_idx}+{c3}{r_idx}, 2)"
                else:
                    formula_ri_clm_incurred = f"=ROUND({curr}{r_idx}-{prev}{r_idx}+{c3}{r_idx}, 2)"
                ws.cell(row = r_idx, column = ri_clm_target_col).value = formula_ri_clm_incurred
            
            for x, y in enumerate(net_clm_incurred_steps):
                net_clm_target_col = 132 + i
                formula_net_clm_incurred = f"={x}{r_idx}-{y}{r_idx}"
                ws.cell(row = r_idx, column = net_clm_target_col).value = formula_net_clm_incurred

            formula_gr_clm_larger_500k = f'=IF(D{r_idx}="IBNR", 0, IF(DI{r_idx}>=500000, 1, 0))'
            formula_gr_clm_larger_1m = f'=IF(D{r_idx}="IBNR", 0, IF(DI{r_idx}>=1000000, 1, 0))'
            formula_event = f"=IFERROR(VLOOKUP(D{r_idx}, EVENT!A:B, 2, 0), 0)"
            formula_status = f'=IF(OR(Z{r_idx} = "", Z{r_idx}=0), "Closed", "Open")'
            formula_valid = f'=IF(D{r_idx}="IBNR", 0, IF(AND(MAX(CR{r_idx}:DI{r_idx})=0, MIN(CR{r_idx}:DI{r_idx})=0), 0, 1))'
            formula_op_yr = f"=YEAR(I{r_idx})"
            formula_product = f'=IF(AND(B{r_idx}="Fac-in", E{r_idx}=43), "MACI", IFERROR(LEFT(F{r_idx}, FIND("/", F{r_idx})-1), 0))'
            formula_ag_classification = f"=VLOOKUP(E{r_idx}, 'Product List'!A:C, 3, 0)"
            ws.cell(row = r_idx, column = 145).value = formula_event
            ws.cell(row = r_idx, column = 144).value = formula_gr_clm_larger_1m
            ws.cell(row = r_idx, column = 143).value = formula_gr_clm_larger_500k
            ws.cell(row = r_idx, column = 142).value = formula_status
            ws.cell(row = r_idx, column = 141).value = formula_product
            ws.cell(row = r_idx, column = 140).value = formula_valid
            ws.cell(row = r_idx, column = 3).value = formula_op_yr
            ws.cell(row = r_idx, column = 1).value = formula_ag_classification
            if row[3] == 'IBNR':
                ws.cell(row = r_idx, column = 141).value = 0
    
    wb['Macau TPL (Gr)']['P82'] = gpe_direct_motor
    print(wb['MACI (Gr,Net)']['P51'].value)
    print(wb['MACI (Gr,Net)']['P52'].value)
    print(wb['MACI (Gr,Net)']['P53'].value)
    wb['MPCC (Gr,RI)']['P84'] = gpe_mpcc
    wb['MACI (Gr,Net)']['P84'] = gpe_maci
    wb['EC Direct']['P82'] = gpe_direct_ec
    wb['EC Fac-in']['P82'] = gpe_fac_in_ec
    # wb.save(output)

# Output report for Claims Triangle
def report_claims_triangle(direct_data, fac_in_data, gross_case_reserve, ri_case_reserve):
    direct_data = direct_data.apply(pd.to_numeric, errors='ignore')
    fac_in_data = fac_in_data.apply(pd.to_numeric, errors='ignore')
    template = r"C:\Users\F170\Desktop\AA Report\Function_Claims_Triangle\template\claims_triangle_template.xlsx"
    output = r"C:\Users\F170\Desktop\AA Report\Function_Claims_Triangle\output\claims_triangle.xlsx"
    
    shutil.copy(template, output)

    wb = load_workbook(output)
    
    # Define your formula logic steps
    gr_clm_incurred_steps = [
            ('K', None, 'AB', 'AS'), ('L', 'K', 'AC', 'AT'), ('M', 'L', 'AD', 'AU'),
            ('N', 'M', 'AE', 'AV'), ('O', 'N', 'AF', 'AW'), ('P', 'O', 'AG', 'AX'),
            ('Q', 'P', 'AH', 'AY'), ('R', 'Q', 'AI', 'AZ'), ('S', 'R', 'AJ', 'BA'),
            ('T', 'S', 'AK', 'BB'), ('U', 'T', 'AL', 'BC'), ('V', 'U', 'AM', 'BD'),
            ('W', 'V', 'AN', 'BE'), ('X', 'W', 'AO', 'BF'), ('Y', 'X', 'AP', 'BG'),
            ('Z', 'Y', 'AQ', 'BH'), ('AA', 'Z', 'AR', 'BI')
    ]

    ri_clm_incurred_steps = [
        ('BJ', None, 'CA'), ('BK', 'BJ', 'CB'), ('BL', 'BK', 'CC'), ('BM', 'BL', 'CD'), ('BN', 'BM', 'CE'), 
        ('BO', 'BN', 'CF'), ('BP', 'BO', 'CG'), ('BQ', 'BP', 'CH'), ('BR', 'BQ', 'CI'), ('BS', 'BR', 'CJ'),
        ('BT', 'BS', 'CK'), ('BU', 'BT', 'CL'), ('BV', 'BU', 'CM'), ('BW', 'BV', 'CN'), ('BX', 'BW', 'CO'),
        ('BY', 'BX', 'CP'), ('BZ', 'BY', 'CQ')
    ]

    net_clm_incurred_steps = [
        ('DB', 'DT'), ('DC', 'DU'), ('DB', 'DT'), ('DD', 'DV'), ('DE', 'DW'), ('DF', 'DX'), ('DG', 'DY'), ('DH', 'DZ')
    ]

    # Map the sheets to their dataframes
    tasks_config = [
        ('Direct Data', direct_data),
        ('Fac-in Data', fac_in_data)
    ]

    for sheet_name, df in tasks_config:
        ws = wb[sheet_name]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)
        
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                if isinstance(value, (int, float)) or value is None:
                    ws.cell(row=r_idx, column=c_idx).value = value
                    continue
                try:
                    ws.cell(row=r_idx, column=c_idx).value = float(value)
                except (ValueError, TypeError):
                    ws.cell(row=r_idx, column=c_idx).value = value
            
            for i, (curr, prev, c3, c4) in enumerate(gr_clm_incurred_steps):
                gr_clm_target_col = 96 + i 
                if prev is None:
                    formula_gr_clm_incurred = f"=ROUND({curr}{r_idx}+{c3}{r_idx}+{c4}{r_idx},2)"
                else:
                    formula_gr_clm_incurred = f"=ROUND({curr}{r_idx}-{prev}{r_idx}+{c3}{r_idx}+{c4}{r_idx},2)"
                
                ws.cell(row=r_idx, column=gr_clm_target_col).value = formula_gr_clm_incurred

            for i, (curr, prev, c3) in enumerate(ri_clm_incurred_steps):
                ri_clm_target_col = 114 + i
                if prev is None:
                    formula_ri_clm_incurred = f"=ROUND({curr}{r_idx}+{c3}{r_idx}, 2)"
                else:
                    formula_ri_clm_incurred = f"=ROUND({curr}{r_idx}-{prev}{r_idx}+{c3}{r_idx}, 2)"
                ws.cell(row = r_idx, column = ri_clm_target_col).value = formula_ri_clm_incurred
            
            for x, y in enumerate(net_clm_incurred_steps):
                net_clm_target_col = 132 + i
                formula_net_clm_incurred = f"={x}{r_idx}-{y}{r_idx}"
                ws.cell(row = r_idx, column = net_clm_target_col).value = formula_net_clm_incurred

            formula_gr_clm_larger_500k = f'=IF(D{r_idx}="IBNR", 0, IF(DI{r_idx}>=500000, 1, 0))'
            formula_gr_clm_larger_1m = f'=IF(D{r_idx}="IBNR", 0, IF(DI{r_idx}>=1000000, 1, 0))'
            formula_event = f"=IFERROR(VLOOKUP(D{r_idx}, EVENT!A:B, 2, 0), 0)"
            formula_status = f'=IF(OR(Z{r_idx} = "", Z{r_idx}=0), "Closed", "Open")'
            formula_valid = f'=IF(D{r_idx}="IBNR", 0, IF(AND(MAX(CR{r_idx}:DI{r_idx})=0, MIN(CR{r_idx}:DI{r_idx})=0), 0, 1))'
            formula_op_yr = f"=YEAR(I{r_idx})"
            formula_product = f'=IF(AND(B{r_idx}="Fac-in", E{r_idx}=43), "MACI", IFERROR(LEFT(F{r_idx}, FIND("/", F{r_idx})-1), 0))'
            formula_ag_classification = f"=VLOOKUP(E{r_idx}, 'Product List'!A:C, 3, 0)"
            ws.cell(row = r_idx, column = 145).value = formula_event
            ws.cell(row = r_idx, column = 144).value = formula_gr_clm_larger_1m
            ws.cell(row = r_idx, column = 143).value = formula_gr_clm_larger_500k
            ws.cell(row = r_idx, column = 142).value = formula_status
            ws.cell(row = r_idx, column = 141).value = formula_product
            ws.cell(row = r_idx, column = 140).value = formula_valid
            ws.cell(row = r_idx, column = 3).value = formula_op_yr
            ws.cell(row = r_idx, column = 1).value = formula_ag_classification
            if row[3] == 'IBNR':
                ws.cell(row = r_idx, column = 141).value = 0
    wb['Movement'].cell(row = 12, column = 18).value = gross_case_reserve
    wb['Movement'].cell(row = 12, column = 37).value = ri_case_reserve
    wb.save(output)
    
# Update direct and fac-in db
def update_db(file_path):
    with pd.ExcelFile(r"C:\Users\F170\Desktop\AA Report\Function_Claims_Triangle\template\claims_triangle_template.xlsx") as xls:
        direct_data_original = pd.read_excel(xls, sheet_name='Direct Data')
        fac_in_data_original = pd.read_excel(xls, sheet_name='Fac-in Data')
    with pd.ExcelFile(file_path) as xls:
        direct_data_new_original = pd.read_excel(file_path, sheet_name = 'Direct-Ceded', skiprows = 1)
        fac_in_data_new_original = pd.read_excel(file_path, sheet_name = 'Accepted-Retro ', skiprows = 1)

    if not os.path.exist('direct_data.db'):
        conn = sqlite3.connect('direct_data.db')


    # direct_data_new_original.columns = TEMPLATE_COLUMNS_CLAIMS_REGISTER_DIRECT
    # fac_in_data_new_original.columns = TEMPLATE_COLUMNS_CLAIMS_REGISTER_FAC_IN

    direct_data_new_original = direct_data_new_original[(direct_data_new_original['Claims NO.'] != 'Clm Mgmt Exp.') & 
                                      (direct_data_new_original['Claims NO.'] != 'IBNR') & 
                                      (direct_data_new_original['Claims NO.'] != 'IBNER') & 
                                      (direct_data_new_original['Claims NO.'] != 'Cat Cover Recovery')]
    direct_data_new_original = direct_data_new_original.dropna(subset = ['Claims NO.'])
    direct_data_w_ibnr = direct_data_original[direct_data_original['Claim No.'] == 'IBNR']
    
    fac_in_data_new_original = fac_in_data_new_original[(fac_in_data_new_original['Claims No.'] != 'Clm Mgmt Exp.') & 
                                      (fac_in_data_new_original['Claims No.'] != 'IBNR') & 
                                      (fac_in_data_new_original['Claims No.'] != 'IBNER')]
    fac_in_data_new_original = fac_in_data_new_original.dropna(subset = ['Claims No.'])
    fac_in_data_w_ibnr = fac_in_data_original[fac_in_data_original['Claim No.'] == 'IBNR']

    for year in years:
        direct_data_new_original = direct_data_new_original.rename(columns = {'Claims NO.': 'Claim No.',
                                                            'Policy NO.': 'Policy No.',
                                                            f'Claim Reserves {year}': f'Claim Reserves \n{year}',
                                                            f'Settled Amount {year}': f'Settled Amount \n{year}',
                                                            f'Excess {year}': f'Excess \n{year}',
                                                            f'Rein. Part. In Claims Reserves {year}': f'RI Claims Reserves \n{year}',
                                                            f'RI Settled Amount {year}': f'RI Settled Amount \n{year}'})
        fac_in_data_new_original = fac_in_data_new_original.rename(columns = {'Claims No.': 'Claim No.',
                                                            f'Claim Reserves {year}': f'Claim Reserves \n{year}',
                                                            f'Settled Amount {year}': f'Settled Amount \n{year}',
                                                            f'Excess {year}': f'Excess \n{year}',
                                                            f'Rein. Part. In Claims Reserves {year}': f'RI Claims Reserves \n{year}',
                                                            f'RI Settled Amount {year}': f'RI Settled Amount \n{year}',
                                                            'Settled Amount (before 2011)': 'Settled Amount \n2010'})


    direct_data_new_original['Date of Occurrence'] = pd.to_datetime(direct_data_new_original['Date of Occurrence'], errors = 'coerce')
    fac_in_data_new_original['Date of Occurrence'] = pd.to_datetime(fac_in_data_new_original['Date of Occurrence'], errors = 'coerce')
    direct_data_original['Date of Occurrence'] = pd.to_datetime(direct_data_original['Date of Occurrence'], errors = 'coerce')
    fac_in_data_original['Date of Occurrence'] = pd.to_datetime(fac_in_data_original['Date of Occurrence'], errors = 'coerce')

    direct_data = direct_data_original[direct_data_original['Date of Occurrence'].dt.year <= 2024]
    direct_data_new = direct_data_new_original[direct_data_new_original['Date of Occurrence'].dt.year > 2024]
    direct_data_new = pd.concat([direct_data_new[:21368], direct_data_new_original[direct_data_new_original['Claim No.'] == 'CLM2026CGPL/0007'], direct_data_new_original[21368:]], ignore_index=True)
    fac_in_data = fac_in_data_original[fac_in_data_original['Date of Occurrence'].dt.year <= 2024]
    fac_in_data_new = fac_in_data_new_original[fac_in_data_new_original['Date of Occurrence'].dt.year > 2024]
    direct_data['Accident yr'] = direct_data['Date of Occurrence'].dt.year
    direct_data_new['Accident yr'] = direct_data_new['Date of Occurrence'].dt.year
    fac_in_data['Accident yr'] = fac_in_data['Date of Occurrence'].dt.year
    fac_in_data_new['Accident yr'] = fac_in_data_new['Date of Occurrence'].dt.year

    direct_merged = pd.concat([direct_data, direct_data_new], ignore_index = True)
    direct_merged['Claim No.'] = direct_merged['Claim No.'].astype(str).str.strip()
    direct_data_new_original['Claim No.'] = direct_data_new_original['Claim No.'].astype(str).str.strip()

    fac_in_merged = pd.concat([fac_in_data, fac_in_data_new, fac_in_data_w_ibnr], ignore_index = True)

    target_cols_direct = [
        'Settled Amount \n2026', 
        'Claim Reserves \n2026', 
        'Excess \n2026',
        'RI Settled Amount \n2026', 
        'RI Claims Reserves \n2026'
    ]

    target_cols_fac_in = [
        'Settled Amount \n2026', 
        'Claim Reserves \n2026', 
        'RI Settled Amount \n2026', 
        'RI Claims Reserves \n2026'
    ]

    direct_merged = direct_merged.merge(
        direct_data_new_original[['Claim No.', 'Type Sap'] + target_cols_direct],
        on=['Claim No.', 'Type Sap'], 
        how='left',
        suffixes=('_original', '_new')
    )


    direct_merged = direct_merged.fillna(0)
    direct_merged.columns = direct_merged.columns.astype(str)
    direct_merged = direct_merged.drop(columns=[col for col in direct_merged.columns if col.endswith('_original')])
    direct_merged = direct_merged.rename(columns={col: col.replace('_new', '') for col in direct_merged.columns if col.endswith('_new')})
    direct_merged['Date of Occurrence'] = pd.to_datetime(direct_merged['Date of Occurrence'], errors='coerce')
    direct_merged = direct_merged.drop_duplicates()
    row_1 = direct_data[(direct_data['Claim No.'] == "Medical Benefit - Mandy's child") & 
                        (direct_data['Date of Occurrence'] == "18/01/2020")].head(1)
    row_2 = direct_data[(direct_data['Claim No.'] == "Medical Benefit - Mandy's child") & 
                        (direct_data['Date of Occurrence'] == "21/09/2020")].head(1)
    row_3 = direct_data[(direct_data['Claim No.'] == "Medical Benefit - Mandy's child") & 
                        (direct_data['Date of Occurrence'] == "18/03/2023")].head(1)
    row_4 = direct_data[(direct_data['Claim No.'] == "CLM2024GMED/0011") & 
                        (direct_data['Settled Amount \n2024'] == 150)].head(1)
    row_5 = direct_data[(direct_data['Claim No.'] == "CLM2024GMED/0011") & 
                        (direct_data['Settled Amount \n2024'] == 300)].head(2)
    row_6 = direct_data[(direct_data['Claim No.'] == "CLM2024GMED/0010") & 
                        (direct_data['Settled Amount \n2024'] == 150)].head(1)
    row_7 = direct_data[(direct_data['Claim No.'] == "CLM2024GMED/0010") & 
                    (direct_data['Settled Amount \n2024'] == 100)].head(1)
    row_8 = direct_data[(direct_data['Claim No.'] == "CLM2024GMED/0010") & 
                    (direct_data['Settled Amount \n2024'] == 144)].head(1)
    row_9 = direct_data[(direct_data['Claim No.'] == "CLM2024GMED/0012") & 
                (direct_data['Settled Amount \n2024'] == 150)].head(3)
    row_10 = direct_data[(direct_data['Claim No.'] == "CLM2024GMED/0012") & 
                (direct_data['Settled Amount \n2024'] == 200)].head(2)
    row_11 = direct_data[(direct_data['Claim No.'] == "CLM2024GMED/0014") & 
                (direct_data['Settled Amount \n2024'] == 100)].head(3)
    row_12 = direct_data[(direct_data['Claim No.'] == "CLM2024GMED/0014") & 
                (direct_data['Settled Amount \n2024'] == 150)].head(3)
    row_13 = direct_data[(direct_data['Claim No.'] == "CLM2024GMED/0014") & 
                (direct_data['Settled Amount \n2024'] == 400)].head(3)
    row_14 = direct_data[(direct_data['Claim No.'] == "CLM2024GMED/0017") & 
                (direct_data['Settled Amount \n2024'] == 150)].head(1)
    row_15 = direct_data[(direct_data['Claim No.'] == "CLM2024GMED/0017") & 
                (direct_data['Settled Amount \n2024'] == 300)].head(1)
    row_17 = direct_data[(direct_data['Claim No.'] == "CLM2024GMED/0018") &
                         (direct_data['Settled Amount \n2025'] == 300)].head(3)
    row_18 = direct_data[(direct_data['Claim No.'] == "CLM2024GMED/0018") &
                         (direct_data['Settled Amount \n2025'] == 1200)].head(1)
    row_19 = direct_data[(direct_data['Claim No.'] == "CLM2024GMED/0018") &
                         (direct_data['Settled Amount \n2025'] == 800)].head(1)
    row_20 = direct_data[(direct_data['Claim No.'] == "CLM2025GMED/0003") &
                         (direct_data['Settled Amount \n2025'] == 150)].head(2)
    row_21 = direct_data[(direct_data['Claim No.'] == "CLM2025GMED/0003") &
                         (direct_data['Settled Amount \n2025'] == 300)].head(1)
    row_22 = direct_data[(direct_data['Claim No.'] == "CLM2025GMED/0003") &
                         (direct_data['Settled Amount \n2025'] == 200)].head(1)
    row_23 = direct_data[(direct_data['Claim No.'] == "CLM2025GMED/0003") &
                         (direct_data['Settled Amount \n2025'] == 400)].head(1)
    row_24 = direct_data[(direct_data['Claim No.'] == "CLM2025GMED/0002") &
                         (direct_data['Settled Amount \n2025'] == 500)].head(2)
    row_25 = direct_data[(direct_data['Claim No.'] == "CLM2025GMED/0002") &
                         (direct_data['Settled Amount \n2025'] == 1200)].head(1)
    row_26 = direct_data[(direct_data['Claim No.'] == "CLM2025GMED/0002") &
                         (direct_data['Settled Amount \n2025'] == 300)].head(1)
    
    direct_merged = pd.concat([direct_merged[:15856], row_1, direct_merged[15856:]], ignore_index=True)
    direct_merged = pd.concat([direct_merged[:16192], row_2, direct_merged[16192:]], ignore_index=True)
    direct_merged = pd.concat([direct_merged[:17995], row_3, direct_merged[17995:]], ignore_index=True)
    direct_merged = pd.concat([direct_merged[:19418], row_4, direct_merged[19418:]], ignore_index=True)
    direct_merged = pd.concat([direct_merged[:19420], row_5, direct_merged[19420:]], ignore_index=True)
    direct_merged = pd.concat([direct_merged[:19457], row_6, row_7, row_8, direct_merged[19457:]], ignore_index=True)
    direct_merged = pd.concat([direct_merged[:19542], row_9, row_10, direct_merged[19542:]], ignore_index=True)
    direct_merged = pd.concat([direct_merged[:19671], row_11, row_12, row_13, direct_merged[19671:]], ignore_index=True)
    direct_merged = pd.concat([direct_merged[:19742], row_14, row_15, direct_merged[19742:]], ignore_index=True)
    direct_merged = pd.concat([direct_merged[:19858], row_17, row_18, row_19, direct_merged[19858:]], ignore_index=True)
    direct_merged = pd.concat([direct_merged[:19940], row_20, row_21, row_22, row_23, direct_merged[19940:]], ignore_index=True)
    direct_merged = pd.concat([direct_merged[:19920], row_24, row_25, row_26, direct_merged[19920:]], ignore_index=True)

    direct_merged = pd.concat([direct_merged, direct_data_w_ibnr], ignore_index = True)


    fac_in_merged = fac_in_merged.merge(
        fac_in_data_new_original[['Claim No.'] + target_cols_fac_in],
        on='Claim No.', 
        how='left',
        suffixes=('_original', '_new')
    )

    for col in target_cols_fac_in:
        # Combine the new and original values
        fac_in_merged[col] = fac_in_merged[f'{col}_new'].combine_first(
            fac_in_merged[f'{col}_original']
        )
        
        # Identify duplicates for this specific Claim No + Amount pair
        is_duplicate = fac_in_merged.duplicated(subset=['Claim No.', col], keep='first')
        
        # Set duplicates to 0
        fac_in_merged.loc[is_duplicate, col] = 0
        
        # Drop the temporary merge columns
        fac_in_merged = fac_in_merged.drop(columns=[f'{col}_original', f'{col}_new'])


    fac_in_merged.loc[fac_in_merged['Claim No.'] == 'CLM2025MACI/0136', 'Settled Amount \n2025'] = 1687.63
    fac_in_merged.loc[fac_in_merged['Claim No.'] == 'CLM2025PAR/0072', 'Settled Amount \n2025'] = 42237

    def prepare_for_sqlite(df):
        for col in df.columns:
            # Case 1: proper datetime dtype
            if pd.api.types.is_datetime64_any_dtype(df[col]):
                df[col] = df[col].dt.strftime('%Y-%m-%d')

            # Case 2: object column that may contain Timestamps
            elif df[col].dtype == object:
                df[col] = df[col].apply(
                    lambda x: x.strftime('%Y-%m-%d')
                    if isinstance(x, pd.Timestamp)
                    else x
                )
        return df

    direct_merged = prepare_for_sqlite(direct_merged)
    fac_in_merged = prepare_for_sqlite(fac_in_merged)
    with pd.ExcelWriter(
        r"C:\Users\F170\Desktop\AA Report\Function_Claims_Triangle\output\output.xlsx", 
        engine='openpyxl'
    ) as writer:
        direct_merged.to_excel(writer, sheet_name='direct', index=False)
        fac_in_merged.to_excel(writer, sheet_name='fac_in', index=False)

    with sqlite3.connect(r"C:\Users\F170\Desktop\AA Report\Function_Claims_Triangle\output\direct_data.db") as conn:
        direct_merged.to_sql('direct_data', conn, if_exists = 'replace', index = False)
    with sqlite3.connect(r"C:\Users\F170\Desktop\AA Report\Function_Claims_Triangle\output\fac_in_data.db") as conn:
        fac_in_merged.to_sql('fac_in_data', conn, if_exists = 'replace', index = False)

# Travel Claims Aaalysis Report
def report_travel_analysis(file):
    months = ['January', 'February', 'March', 'April', 'May' , 'June', 'July', 'August', 'September', 'October','November', 'December']
    all_monthly_dfs = []

    for year in years:
        sheet_name = f"{year}Travel"
        df = pd.read_excel(file, sheet_name = sheet_name, engine = 'xlrd')
        df['Reported Date'] = pd.to_datetime(df['Reported Date'], errors = 'coerce')
        df = df[df['Reported Date'].dt.year == year]

        df['Year'] = df['Reported Date'].dt.year
        df['Month'] = df['Reported Date'].dt.month
        df['Month_Name'] = df['Reported Date'].dt.strftime('%B')
        all_monthly_dfs.append(df)

    df_all = pd.concat(all_monthly_dfs, ignore_index = True)

    df_all['Claimed Policy Section'] = (
        df_all['Claimed Policy Section']
        .astype(str)                         
        .str.strip()                        
        .str.replace(r'\s*\n\s*', ' & ', regex=True)   
        .str.replace(r'\s+', ' ', regex=True)        
        .str.strip()
    )

    # Mapping must be after the replace string
    mapping = {
        'Rental Vehicle Excess': 'rental car',
        'Medical expenses & & follow-up medical treatment': 'medical expenses',
        'Luggage Damaged': 'personal belongings',
        'Trip Curtailmant': 'travel cancellation',
        'Special Care': 'other',
        'Trip Cancellation': 'travel cancellation',
        'Personal Money': 'personal belongings',
        'Personal Baggage': 'personal belongings',
        'trip re-arrangement': 'travel cancellation',
        'Personal Liability': 'personal belongings',
        'the i-pad2 lost on the airline': 'personal belongings',
        'wallet lost': 'personal belongings',
        'the phone fall down in lake': 'personal belongings',
        'Loss of I - pad3': 'personal belongings',
        'Loss of money': 'personal belongings',
        'loss of phone': 'personal belongings',
        'sudden loss of eyesight on left eye': 'medical expenses',
        'personal money': 'personal belongings',
        'Belonging damaged': 'personal belongings',
        'Loss of travel document': 'personal belongings',
        'Camera Len was damaged': 'personal belongings',
        'Damaged rental car during the trip': 'rental car',
        'The luggage case was damage while they picked it up': 'baggage damaged',
        'travel delay & baggage damage': 'travel delay', 
        'Baggage delay & damage': 'baggage damaged',
        'Travel delay & baggage delay': 'travel delay',
        'Rental vehicle excess': 'rental car',
        'Travel delay & Mecical expenses': 'medical expenses',
        'Medical expense': 'medical expenses',
        'Baggage damaged & Medical expenses': 'medical expenses',
        'Issue payment': 'medical expenses', 
        'Medical expenses & Personal belongings': 'medical expenses',
        'Personal Money & Loss of Travel document': 'personal belongings',
        'Personal belongings & Trip Cancellation': 'travel cancellation',
        'Personal Money & Personal Belonging': 'personal belongings',
        'Lost of travel document': 'personal belongings',
        'Bagagge damaged': 'baggage damaged',
        'Personal belongings & Personal Money & Loss of Travel document': 'personal belongings',
        'Rental Vehicle Excess': 'rental car',
        'Medical expenses & Loss of travel document': 'medical expenses',
        'Personal belongings & Personal Money': 'personal belongings',
        'Rental Vehicle Excess & Travel delay': 'rental car',
        'Medical expenses & Overseas Hospital Cash': 'medical expenses',
        'Baggage delay & Personal belongings': 'baggage delay',
        'Meical expenses': 'medical expenses',
        'Bagagge delay': 'baggage delay',
        'Loss of travel documents': 'personal belongings',
        'Baggage damaged & Rental vehicle excess': 'Baggage damaged',
        'Trave delay & Personal belongings': 'personal belongings',
        'Medical expenses & Trip curtailment': 'medical expenses',
        'Medical expenses & Overseas hospital cash': 'medical expenses',
        'Personal belongings & Baggage damgaed': 'personal belongings',
        'Personal money & Loss of Travel document & Emergency Cash': 'personal belongings',
        'Unoccupied household burglary': 'personal belongings',
        'Medical expenses & Personal belongings & Rental vehicle excess': 'Medical expenses',
        'Personal belongings & Personal money & Loss of travel document': 'personal belongings',
        'Personal belongigns': 'personal belongings',
        'Loss of Travel document': 'personal belongings',
        'Trip re-arrangement & Travel delay': 'travel delay',
        'Medical expenses & Trip cancellation & Trip curtailment': 'travel cancellation',
        'Trip re-arrangement & Baggage damaged': 'travel re-arrangement', 
        'Trip re-arrangement & Baggage delay': 'travel re-arrangement',
        'Trip re-arrangemant': 'travel re-arrangement',
        'Medical expenses & Baggage delay': 'medical expenses',
        'Personal belongings & Personal money & Loss of travel documents': 'personal belongings',
        'Trip cancellation & Travel delay': 'travel delay', 
        'Rental vehicle excess & Personal liability': 'rental car',
        'Personal belongings & Loss of travel documents & Personal money': 'personal belongings',
        'Personal beloingings & Loss of Travel documents': 'personal belongings',
        'Tavel delay': 'travel delay',
        'Personal belonings': 'personal belongings',
        'Personal money & Loss of travel document': 'personal belongings',
        'Mediacal expenses & Travel delay': 'medical expenses',
        'Medical expenses & Overseas Hospital Cash & Trip re-arrangement': 'medical expenses',
        'Baggage Delay': 'baggage delay',
        'Trip curtailment & Medical Expenses': 'medical expenses',
        'Mediacal expeneses': 'medical expenses',
        'Other accident': 'other',
        'Re-arrangement': 'travel re-arrangement',
        'Medical expenses & Personal baggage & Personal money & Loss of travel document & Emergency cash': 'personal belongings',
        '': 'unknown',
        'Baggage delay & & Personal belongings': 'baggage delay',
        'Travel delay& Special & care': 'travel delay',
        'Personal belongings,personal cash': 'personal belongings',
        'Medical expenses& Baggage damage': 'personal belongings',
        'Medical expenses&Baggage damage': 'medical expenses',
        'Travel delay&Baggage damage': 'baggage damaged',
        'Personal belongings & Special care vehicle excess': 'personal belongings',
        'Personal money & Loss of Travel document': 'personal belongings',
        'Personal baggage & personal money': 'personal belongings',
        'Medical Expenses Related Costs': 'medical expenses',
        'Medical expenses and follow up treatment': 'medical expenses',
        'Travel Delay': 'travel delay', 
        'Travel delay & Personal Belongings': 'travel delay',
        'Travel delay & Baggage damage': 'baggage damaged',
        'Travel delay / Medical expenses / Personal Belongings': 'travel delay',
        'Travel delay / Medical expenses': 'travel delay',
        'Medical Expenses / Related Costs': 'medical expenses',
        'Trip Re-routing (Re-arrangement)': 'travel re-arrangement',
        'Trip curtailment / Medical expenses & follow up treatment': 'medical expenses',
        'Trip Re-routing & (Re-arrangement)': 'travel re-arrangement',
        'Travel delay & Medical expenses': 'medical expenses',
        'Trip Curtailment / Re-routing (Re-arrangement)': 'travel re-arrangement',
        'Medical expenses / Trip Curtailment': 'medical expenses',
        'Travel delay / Baggage delay and Personal belongings': 'personal belongings',
        'Personal belongings / Medical expenses & follow-up medical treatment': 'medical expenses',
        'travel delay / baggage delay': 'travel delay',
        'Personal Belongings': 'personal belongings',
        'Travel delay': 'travel delay',
        'Medical expenses': 'medical expenses',
        'Luggage damaged': 'baggage damaged',
        'Baggage delay': 'baggage delay',
        'Baggage damaged': 'baggage damaged',
        'Trip Re-arrangement': 'travel re-arrangement',
        'Medical expsnses': 'medical expenses',
        'Loss of belonging': 'personal belongings',
        'phone and camera damaged as wet by seawater': 'personal belongings',
        'Baggage delay / Baggage damaged': 'baggage damaged',
        'loss of Ipad': 'personal belongings',
        'medical expnses': 'medical expenses',
        'Medical expenses & Trip re-arrangement': 'travel re-arrangement',
        'Medical expenses & Travel delay': 'medical expenses',
        'Baggage damaged & Baggage delay': 'baggage damaged',
        'Personal money': 'personal belongings',
        'Trip curtailment': 'travel cancellation',
        'Travel delay & Baggage delay': 'baggage delay',
        'Travel delay & Baggage damaged': 'baggage damaged',
        'Personal belongings & Personal money': 'personal belongings',
        'Travel delay & Personal belongings': 'personal belongings',
        'Medical expenses & Baggage damaged': 'medical expenses',
        'Baggage delay & Baggage damaged': 'baggage damaged',
        'Medical expenses & Personal money & Loss of Travel document': 'personal belongings',
        'Baggage damaged & Travel delay': 'baggage damaged',
        'Personal money & Loss of travel documents': 'personal belongings',
        'Baggage damged': 'baggage damaged',
        'Medical expenses & Trip cancellation': 'medical expenses',
        'Trip re-arrangement & Travel delay & Rental vehicle excess': 'Trip re-arrangementr',
        'Personal belongins': 'personal belongings',
        'travel canceallation': 'travel cancellation',
        'Travel delay & Rental Vehicle Excess': 'Travel delay',
        'Personal baggage & & Personal money& & Loss of travel document': 'personal belongings',
        'Travel delay& Baggage delay': 'baggage delay',
        'Follow-up medical treatment': 'medical expenses',
        'Baggage delay & & Baggage damaged': 'baggage damaged',
        'Medical expenses&Follow-up medical treatment': 'medical expenses',
        'Personal money & & Loss of travel document': 'personal belongings',
        'Medical expenses& & follow up medical treatment': 'medical expenses',
        'Medical expenses & & follow up medical treatment': 'medical expenses',
        'Baggage delay& & Personal belongings': 'personal belongings',
        'Travel delay & & Baggage delay': 'baggage delay',
        'Medical expenses & &Travel delay': 'medical expenses',
        'Travel delay&Baggage delay': 'baggage delay',
        'Baggage damage': 'baggage damaged',
        'Personal baggage & Baggage delay': 'personal belongings',
        'Medical expenses & & Trip Re-arrangement': 'medical expenses',
        'Personal Baggage & Personal Money': 'personal belongings',
        'Personal baggage & Personal money & Loss of travel document': 'personal belongings',
        'Medical expenses & follow up treatment': 'medical expenses',
        'Trip cancellation & baggage damage': 'baggage damaged',
        'Baggage delay&Baggage damage': 'baggage damaged',
        'Special care vehicle excess': 'other',
        'Medical expenses & Follow up treatment': 'medical expenses',
        'Medical expenses%follow up treatment': 'medical expenses',
        'Trip Cancellation / Curtailment / Re-arrangement/Personal belongings': 'personal belongings',
        'Medical expenses&follow up treatment & Overseas hospital cash': 'medical expenses',
        'Baggage demaged': 'baggage damaged',
        'Personal belongings& Baggage demaged': 'personal belongings',
        'Travel delay& Baggage damage': 'baggage damaged',
        'baggage damage': 'baggage damaged',
        'Travel delay&Baggage damaged': 'baggage damaged',
        'Medical Expenses': 'medical expenses',
        'Travel delay / Baggage damaged / Baggage delay': 'baggage damaged',
        'Trip cancellation': 'travel cancellation',
        'Trip re-arrangement': 'travel re-arrangement',
        'Medical expenses & Personal baggage': 'medical expenses',
        'Personal baggage': 'personal belongings',
        'medical expnse': 'medical expenses',
        'Medical expenses & & Trip re-arrangement': 'medical expenses',
        'Personal baggage& & Personal money & & Loss of travel document': 'personal belongings',
        'Trip re-arrangement& & baggage damage': 'baggage damaged',
        'Trip Curtailment': 'travel cancellation',
        'Travel delay& & Medical expenses': 'medical expenses',
        'Travel delay & & baggage damage': 'baggage damaged',
        'Medical expenses & & Trip Curtailment': 'medical expenses',
        'Medical expenses & Trip Curtailment & & Personal belongings': 'medical expenses',
        'Personal Money &Loss of travel document': 'personal belongings',
        'Medical expenses&follow-up medical treatment': 'medical expenses',
        'travel delay / personal belongings': 'personal belongings',
        'personal Money': 'personal belongings',
        'Personal belongings': 'personal belongings',
        'Medical expenses (follow-up medical treatment)': 'medical expenses'
        
    }

    df_all['Claimed Policy Section'] = df_all['Claimed Policy Section'].replace(mapping)
    df_all['Year'] = df_all['Reported Date'].dt.year
    df_all['Month'] = df_all['Reported Date'].dt.month
    df_all['Month_Name'] = df_all['Month'].apply(lambda x: months[x-1])
    df_all['Total paid (MOP)'] = pd.to_numeric(df_all['Total paid (MOP)'], errors='coerce')
    df_all['Reserve'] = pd.to_numeric(df_all['Reserve'], errors='coerce').fillna(0)
    df_all['Paid Medical'] = pd.to_numeric(df_all['Paid Medical'], errors='coerce').fillna(0)
    df_all['Paid Others'] = pd.to_numeric(df_all['Paid Others'], errors='coerce').fillna(0)
    df_all['Total paid'] = df_all['Paid Medical'] + df_all['Paid Others'] + df_all['Reserve']
    all_categories = df_all['Claimed Policy Section'].dropna().unique()
    all_categories = sorted(all_categories)

    # Export to Excel
    wb = load_workbook(r"C:\Users\F170\Desktop\AA Report\Function_Travel_Analysis\template\template.xlsx")

    ws_raw_data = wb['Raw Data']
    rows = dataframe_to_rows(df_all, index=False, header=True)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws_raw_data.cell(row=r_idx, column=c_idx, value=value)

    timestamp = datetime.now().strftime("%Y-%m-%d")
    wb.save(rf"C:\Users\F170\Desktop\AA Report\Function_Travel_Analysis\output\BNU_Travel_{timestamp}.xlsx")

def checking_claim_triangle(direct_data, fac_in_data, file_path):
    direct_data_updated = direct_data[direct_data['Accident yr'] == max(years)]
    fac_in_data_updated = fac_in_data[fac_in_data['Accident yr'] == max(years)]
    with pd.ExcelFile(file_path) as xls:
        direct_data_new = pd.read_excel(xls, sheet_name = 'Direct-Ceded', skiprows = 1, 
                                        usecols = ['Date of Occurrence', 'Settled Amount 2026', 'Claim Reserves 2026', 'Excess 2026', 
                                                   'RI Settled Amount 2026', 'Rein. Part. In Claims Reserves 2026'])
        fac_in_data_new = pd.read_excel(xls, sheet_name = 'Accepted-Retro ', skiprows = 1, 
                                        usecols = ['Date of Occurrence', 'Settled Amount 2026', 'Claim Reserves 2026', 
                                                   'RI Settled Amount 2026', 'Rein. Part. In Claims Reserves 2026'])
        direct_data_new['Date of Occurrence'] = pd.to_datetime(direct_data_new['Date of Occurrence'], errors = 'coerce')
        fac_in_data_new['Date of Occurrence'] = pd.to_datetime(fac_in_data_new['Date of Occurrence'], errors = 'coerce')
        direct_data_new = direct_data_new[direct_data_new['Date of Occurrence'].dt.year == max(years)]
        fac_in_data_new = fac_in_data_new[fac_in_data_new['Date of Occurrence'].dt.year == max(years)]
        
    # Direct
    claim_reserves_direct_new = direct_data_new['Claim Reserves 2026'].sum()
    settled_amount_direct_new = direct_data_new['Settled Amount 2026'].sum()
    excess_direct_new = direct_data_new['Excess 2026'].sum()
    ri_settled_amount_direct_new = direct_data_new['RI Settled Amount 2026'].sum()
    ri_claim_reserves_direct_new = direct_data_new['Rein. Part. In Claims Reserves 2026'].sum()

    claim_reserves_direct_updated = direct_data_updated['Claim Reserves \n2026'].sum()
    settled_amount_direct_updated = direct_data_updated['Settled Amount \n2026'].sum()
    excess_direct_updated = direct_data_updated['Excess \n2026'].sum()
    ri_settled_amount_direct_updated = direct_data_updated['RI Settled Amount \n2026'].sum()
    ri_claim_reserves_direct_updated = direct_data_updated['RI Claims Reserves \n2026'].sum()

    # Fac-in
    claim_reserves_fac_in_new = fac_in_data_new['Claim Reserves 2026'].sum()
    settled_amount_fac_in_new = fac_in_data_new['Settled Amount 2026'].sum()
    ri_settled_amount_fac_in_new = fac_in_data_new['RI Settled Amount 2026'].sum()
    ri_claim_reserves_fac_in_new = fac_in_data_new['Rein. Part. In Claims Reserves 2026'].sum()

    claim_reserves_fac_in_updated = fac_in_data_updated['Claim Reserves \n2026'].sum()
    settled_amount_fac_in_updated = fac_in_data_updated['Settled Amount \n2026'].sum()
    ri_settled_amount_fac_in_updated = fac_in_data_updated['RI Settled Amount \n2026'].sum()
    ri_claim_reserves_fac_in_updated = fac_in_data_updated['RI Claims Reserves \n2026'].sum()


    rows_direct = [
        'Claim Reserves', 'Settled Amount', 'Excess', 
        'RI Settled Amount', 'RI Claim Reserves'
    ]

    data_direct_values = {
        'New': [claim_reserves_direct_new, settled_amount_direct_new, excess_direct_new, 
                ri_settled_amount_direct_new, ri_claim_reserves_direct_new],
        'Update': [claim_reserves_direct_updated, settled_amount_direct_updated, excess_direct_updated, 
                ri_settled_amount_direct_updated, ri_claim_reserves_direct_updated]
    }

    df_direct = pd.DataFrame(data_direct_values, index=rows_direct)

    # Calculate Differences and Totals
    df_direct['Differences'] = (df_direct['New'] - df_direct['Update']).round(0)
    df_direct.loc['Total'] = df_direct.sum()

    rows_fac_in = [
        'Claim Reserves', 'Settled Amount', 
        'RI Settled Amount', 'RI Claim Reserves'
    ]

    data_fac_in_values = {
        'New': [claim_reserves_fac_in_new, settled_amount_fac_in_new, 
                ri_settled_amount_fac_in_new, ri_claim_reserves_fac_in_new],
        'Update': [claim_reserves_fac_in_updated, settled_amount_fac_in_updated, 
                ri_settled_amount_fac_in_updated, ri_claim_reserves_fac_in_updated]
    }

    df_fac_in = pd.DataFrame(data_fac_in_values, index=rows_fac_in)

    # Calculate Differences and Totals
    df_fac_in['Differences'] = (df_fac_in['New'] - df_fac_in['Update']).round(0)
    df_fac_in.loc['Total'] = df_fac_in.sum()

    return df_direct, df_fac_in

# report_output_ey(options = 'Fire')
# update_db(r"C:\Users\F170\Desktop\AA Report\Function_Claims_Triangle\template\claims_register_new.xlsx")
# report_claims_triangle(direct_data, fac_in_data, gross_case_reserve = 42098246, ri_case_reserve= 21552062.42)
# report_travel_analysis(r"C:\Users\F170\Desktop\AA Report\Function_Travel_Analysis\data\BNU Travel Register _ 04.xls")
report_claims_reserve()
# checking_claim_triangle(direct_data, fac_in_data, r"C:\Users\F170\Desktop\AA Report\Function_Claims_Triangle\template\claims_register_new.xlsx")




